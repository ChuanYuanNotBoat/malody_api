from __future__ import annotations

import pandas as pd


def is_change_column(column_name: str) -> bool:
    normalized = column_name.strip().lower()
    keywords = ("change", "delta", "diff", "变化", "变动", "涨跌", "_chg", "chg_")
    return any(keyword in normalized for keyword in keywords)


def is_rank_change_column(column_name: str) -> bool:
    normalized = column_name.strip().lower()
    return is_change_column(column_name) and ("rank" in normalized or "排名" in normalized)


def autosize_openpyxl_sheet(worksheet) -> None:
    try:
        for column_cells in worksheet.columns:
            max_len = 0
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_len:
                    max_len = len(value)
            column_letter = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = min(max(max_len + 2, 10), 60)
    except Exception:
        # Non-openpyxl worksheet engines may not expose worksheet.columns/column_dimensions.
        return


def apply_change_conditional_formatting(worksheet, dataframe: pd.DataFrame) -> None:
    # OpenPyXL-only enhancement: conditionally color "change" columns.
    try:
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return

    if dataframe.empty:
        return

    # Use ARGB + start/end colors for broader Excel compatibility in conditional formats.
    red_fill = PatternFill(fill_type="solid", start_color="FFFCE4D6", end_color="FFFCE4D6")
    red_font = Font(color="FF9C0006")
    green_fill = PatternFill(fill_type="solid", start_color="FFE2F0D9", end_color="FFE2F0D9")
    green_font = Font(color="FF006100")

    for idx, column_name in enumerate(dataframe.columns, start=1):
        if not is_change_column(str(column_name)):
            continue
        if not pd.api.types.is_numeric_dtype(dataframe[column_name]):
            continue

        column_letter = get_column_letter(idx)
        data_range = f"{column_letter}2:{column_letter}{len(dataframe) + 1}"
        rank_change = is_rank_change_column(str(column_name))

        if rank_change:
            # Rank number down (<0) means rank up; apply red style per spec.
            worksheet.conditional_formatting.add(
                data_range,
                CellIsRule(operator="lessThan", formula=["0"], stopIfTrue=False, fill=red_fill, font=red_font),
            )
            worksheet.conditional_formatting.add(
                data_range,
                CellIsRule(operator="greaterThan", formula=["0"], stopIfTrue=False, fill=green_fill, font=green_font),
            )
        else:
            worksheet.conditional_formatting.add(
                data_range,
                CellIsRule(operator="greaterThan", formula=["0"], stopIfTrue=False, fill=red_fill, font=red_font),
            )
            worksheet.conditional_formatting.add(
                data_range,
                CellIsRule(operator="lessThan", formula=["0"], stopIfTrue=False, fill=green_fill, font=green_font),
            )
