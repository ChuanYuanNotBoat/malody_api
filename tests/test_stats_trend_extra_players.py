import sqlite3
import sys
from datetime import datetime
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from types import SimpleNamespace
from unittest import TestCase
from unittest.mock import patch

from openpyxl import load_workbook

ROOT_PARENT = Path(__file__).resolve().parents[2]
if str(ROOT_PARENT) not in sys.path:
    sys.path.insert(0, str(ROOT_PARENT))

from malody_api.stats_cli.plugins import trend_plugin  # noqa: E402


def _colorize(text: str, _color: str) -> str:
    return str(text)


def _db_safe_operation(func):
    def wrapper(self, *args, **kwargs):
        return func(self, *args, **kwargs)

    wrapper.__doc__ = getattr(func, "__doc__", None)
    return wrapper


def _separator(_width=None) -> str:
    return "----"


class _TrendShellStub:
    def __init__(self, conn, output_dir):
        self.conn = conn
        self.output_dir = output_dir
        self.current_mode = 3
        self.mode_names = {-1: "All", 3: "Catch"}
        self.selector = SimpleNamespace(get_current_selection=lambda: "Mode: 3(Catch)")
        self._non_interactive = True

    def get_unique_filename(self, base, ext):
        return f"{Path(base).stem}.{ext}"


class TestStatsTrendExtraPlayers(TestCase):
    @classmethod
    def setUpClass(cls):
        colors = SimpleNamespace(RED="RED", YELLOW="YELLOW", GREEN="GREEN", CYAN="CYAN", BLUE="BLUE", BOLD="BOLD")
        trend_plugin.install(
            _TrendShellStub,
            colorize=_colorize,
            colors=colors,
            db_safe_operation=_db_safe_operation,
            get_separator=_separator,
        )

    def setUp(self):
        self.tempdir = TemporaryDirectory()
        self.conn = sqlite3.connect(":memory:", detect_types=sqlite3.PARSE_DECLTYPES)
        cursor = self.conn.cursor()
        cursor.executescript(
            """
            CREATE TABLE player_rankings (
                player_id INTEGER,
                name TEXT,
                rank INTEGER,
                lv INTEGER,
                exp INTEGER,
                acc REAL,
                combo INTEGER,
                pc INTEGER,
                mode INTEGER,
                crawl_time timestamp
            );
            CREATE TABLE player_identity (
                player_id INTEGER,
                uid INTEGER,
                current_name TEXT
            );
            CREATE TABLE player_aliases (
                player_id INTEGER,
                alias TEXT
            );
            """
        )
        cursor.executemany(
            """
            INSERT INTO player_rankings
            (player_id, name, rank, lv, exp, acc, combo, pc, mode, crawl_time)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (1, "TopA", 5, 10, 1000, 95.0, 100, 10, 3, datetime(2026, 5, 1, 0, 0, 0)),
                (1, "TopA", 4, 11, 1100, 95.5, 120, 11, 3, datetime(2026, 5, 7, 0, 0, 0)),
                (2, "TrackedB", 60, 9, 900, 93.0, 80, 8, 3, datetime(2026, 5, 1, 0, 0, 0)),
                (2, "TrackedB", 58, 10, 980, 93.3, 90, 9, 3, datetime(2026, 5, 7, 0, 0, 0)),
            ],
        )
        cursor.executemany(
            "INSERT INTO player_identity (player_id, uid, current_name) VALUES (?, ?, ?)",
            [
                (1, 1001, "TopA"),
                (2, 2002, "TrackedB"),
            ],
        )
        self.conn.commit()
        self.shell = _TrendShellStub(self.conn, self.tempdir.name)

    def tearDown(self):
        self.conn.close()
        self.tempdir.cleanup()

    def test_do_trend_can_include_tracked_players_outside_rank_range_and_export_non_interactive(self):
        with patch("malody_api.stats_cli.plugins.trend_plugin.get_terminal_width", return_value=400), patch(
            "malody_api.stats_cli.plugins.trend_plugin._load_tracked_tokens", return_value=["2002"]
        ):
            self.shell.do_trend(
                "2026-05-01 --mode 3 --rank-range 1-50 --include-tracked --show-uid --export-uid --export xlsx"
            )

        workbook_bytes = Path(self.tempdir.name, "trend_mode3_20260501_20260507.xlsx").read_bytes()
        workbook = load_workbook(BytesIO(workbook_bytes), read_only=True)
        sheet = workbook["trend"]
        rows = list(sheet.iter_rows(values_only=True))
        headers = rows[0]
        names = [r[1] for r in rows[1:] if r and len(r) > 1]

        self.assertIn("TopA", names)
        self.assertIn("TrackedB", names)
        self.assertEqual(rows[1][1], "TrackedB")
        self.assertIn("uid", headers)
        self.assertNotIn("来源Top范围", headers)
        self.assertNotIn("来源额外玩家", headers)
        del sheet
        workbook.close()
        del workbook
