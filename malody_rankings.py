import argparse
import gc
import json
import logging
import os
import queue
import re
import signal
import sqlite3
import subprocess
import sys
import threading
import time
from datetime import datetime, timedelta
from threading import Lock
from typing import Any, Dict, List, Optional, Set, Tuple

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# Color support (enabled only when terminal supports ANSI).
USE_COLOR = hasattr(sys.stdout, 'isatty') and sys.stdout.isatty()
def colorize(text, color_code):
    """Add ANSI color escape sequence when terminal supports color."""
    if USE_COLOR:
        return f"\033[{color_code}m{text}\033[0m"
    return text


def notify_user(title: str, message: str, level: str = "warning"):
    """
    Show a non-blocking Windows toast notification.
    Used to alert users when crawler fails to fetch data,
    especially useful for scheduled tasks where the terminal
    auto-closes after execution.

    Args:
        title: Notification title (e.g. "排行榜爬取异常")
        message: Notification body with failure details
        level: "warning" or "error" (currently unused, reserved for future styling)
    """
    if sys.platform != "win32":
        logger.info("[通知] %s: %s", title, message)
        return

    # Escape XML special characters for the toast template.
    import html as _html_mod
    def _xml_escape(s: str) -> str:
        return _html_mod.escape(s, quote=True)
    safe_title = _xml_escape(title)
    safe_message = _xml_escape(message)

    ps_script = (
        '[Windows.UI.Notifications.ToastNotificationManager, Windows.UI.Notifications, '
        'ContentType = WindowsRuntime] | Out-Null\n'
        '[Windows.Data.Xml.Dom.XmlDocument, Windows.Data.Xml.Dom.XmlDocument, '
        'ContentType = WindowsRuntime] | Out-Null\n'
        '$template = @"\n'
        '<toast duration="long">\n'
        '  <visual>\n'
        '    <binding template="ToastGeneric">\n'
        f'      <text>{safe_title}</text>\n'
        f'      <text>{safe_message}</text>\n'
        '    </binding>\n'
        '  </visual>\n'
        '</toast>\n'
        '"@\n'
        '$xml = New-Object Windows.Data.Xml.Dom.XmlDocument\n'
        '$xml.LoadXml($template)\n'
        '$toast = [Windows.UI.Notifications.ToastNotification]::new($xml)\n'
        '[Windows.UI.Notifications.ToastNotificationManager]::CreateToastNotifier("Malody Rankings Crawler").Show($toast)\n'
    )

    try:
        subprocess.Popen(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps_script],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
        logger.info("已发送桌面通知: %s - %s", title, message)
    except Exception as e:
        logger.warning("发送桌面通知失败: %s", e)

# Silence Python 3.12+ SQLite datetime adapter deprecation warning.
def adapt_datetime(dt):
    return dt.isoformat()

def convert_datetime(s):
    return datetime.fromisoformat(s.decode())

sqlite3.register_adapter(datetime, adapt_datetime)
sqlite3.register_converter("timestamp", convert_datetime)

try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False
    print("tqdm not installed; using simple progress output.")

# Logging configuration.
logging.basicConfig(
    filename='crawler.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    filemode='a'
)
logger = logging.getLogger()
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
logger.addHandler(console_handler)

# Cookie configuration (update with real values before running).
def _load_cookies() -> dict:
    """
    Load cookies from runtime sources, highest priority first:
    1) MALODY_COOKIES_JSON (full JSON string)
    2) MALODY_COOKIES_FILE (json file path, default: cookies.local.json)
    3) MALODY_SESSIONID / MALODY_CSRFTOKEN
    
    Supports two cookie file formats:
    - Dict format: {"sessionid": "...", "csrftoken": "..."}
    - Array format: [{"name": "sessionid", "value": "..."}, ...]
    """
    def _extract_cookies_from_data(data):
        """Extract cookies from dict or array format."""
        if isinstance(data, dict):
            return data
        elif isinstance(data, list):
            # Convert array format [{"name": "...", "value": "..."}, ...] to dict
            result = {}
            for item in data:
                if isinstance(item, dict) and "name" in item and "value" in item:
                    result[item["name"]] = item["value"]
            return result
        return {}
    
    cookies = {}

    cookies_file = os.getenv("MALODY_COOKIES_FILE", "cookies.local.json")
    if os.path.exists(cookies_file):
        try:
            with open(cookies_file, "r", encoding="utf-8") as f:
                file_cookies = json.load(f)
            extracted = _extract_cookies_from_data(file_cookies)
            if extracted:
                cookies.update(extracted)
                logger.info("Loaded crawler cookies from %s", cookies_file)
        except Exception as e:
            logger.warning("Failed to load cookies file %s: %s", cookies_file, e)

    cookies_json = os.getenv("MALODY_COOKIES_JSON")
    if cookies_json:
        try:
            env_cookies = json.loads(cookies_json)
            extracted = _extract_cookies_from_data(env_cookies)
            if extracted:
                cookies.update(extracted)
                logger.info("Loaded crawler cookies from MALODY_COOKIES_JSON")
        except Exception as e:
            logger.warning("Invalid MALODY_COOKIES_JSON: %s", e)

    sessionid = os.getenv("MALODY_SESSIONID")
    csrftoken = os.getenv("MALODY_CSRFTOKEN")
    if sessionid:
        cookies["sessionid"] = sessionid
    if csrftoken:
        cookies["csrftoken"] = csrftoken

    if not cookies.get("sessionid"):
        logger.warning("No sessionid configured for crawler cookies.")
    if not cookies.get("csrftoken"):
        logger.warning("No csrftoken configured for crawler cookies.")

    return cookies


COOKIES = _load_cookies()

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Referer": "https://m.mugzone.net/",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
}

BASE_URL = "https://m.mugzone.net/page/all/player?from=0&mode={mode}"
PLAYER_PROFILE_URL = "https://m.mugzone.net/accounts/user/{player_id}"
MODES = list(range(10))
NEW_API_BASE_URL = "https://api.mugzone.net/api"
NEW_API_MM_PAGE_SIZE = 40
DEFAULT_MM_LIMIT = 200
MM_RUN_LOCK_FILE = "tmp/mm_sync.lock"

DB_FILE = "malody_rankings.db"
GIT_REPO_PATH = os.path.dirname(os.path.abspath(__file__))
GIT_COMMIT_MESSAGE = datetime.now().strftime("%Y-%m-%d %H:%M updated")

stop_requested = False
skip_current_requested = False
last_sigint_time = 0.0
SIGINT_STOP_WINDOW_SECONDS = 2.0
stop_lock = Lock()

# Player configuration file.
PLAYER_CONFIG_FILE = "players.txt"

# Player crawl queue and state (with global dedupe set).
player_queue = queue.Queue()
_player_set = set()          # 用于去重
_player_set_lock = Lock()
player_crawl_lock = Lock()
player_crawl_in_progress = False
last_player_crawl_time = None

_api_auth_lock = Lock()
_api_auth_cache: Dict[str, Any] = {}

class SkipCurrentTask(Exception):
    """Soft interrupt: skip current crawl unit and continue."""


def is_stop_requested() -> bool:
    with stop_lock:
        return stop_requested


def consume_skip_request() -> bool:
    global skip_current_requested
    with stop_lock:
        if skip_current_requested:
            skip_current_requested = False
            return True
        return False


def raise_if_skip_requested(context: str = "current task"):
    if consume_skip_request():
        raise SkipCurrentTask(context)


def signal_handler(sig, frame):
    """Handle signals with two-stage Ctrl+C behavior."""
    global stop_requested, skip_current_requested, last_sigint_time
    with stop_lock:
        if sig == signal.SIGTERM:
            stop_requested = True
            logger.warning("Received SIGTERM, stopping crawler safely...")
            return

        now = time.time()
        if (now - last_sigint_time) <= SIGINT_STOP_WINDOW_SECONDS:
            stop_requested = True
            logger.warning("Received Ctrl+C twice quickly: stopping crawler safely...")
        else:
            skip_current_requested = True
            logger.warning("Received Ctrl+C: skip current crawl item/source and continue.")
            logger.warning(
                "Press Ctrl+C again within %.0f seconds to stop the crawler safely.",
                SIGINT_STOP_WINDOW_SECONDS,
            )
        last_sigint_time = now

signal.signal(signal.SIGINT, signal_handler)
signal.signal(signal.SIGTERM, signal_handler)

def get_git_commit_message():
    """Build the Git commit message."""
    return datetime.now().strftime("%Y-%m-%d %H:%M updated")


class DatabaseManager:
    """Database connection manager with per-thread connections."""
    _instance = None
    _lock = Lock()

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super(DatabaseManager, cls).__new__(cls)
            cls._instance.connections = {}
        return cls._instance

    def get_connection(self, thread_id=None):
        """Get the database connection for current thread."""
        if thread_id is None:
            thread_id = threading.get_ident()
        with self._lock:
            if thread_id not in self.connections:
                self.connections[thread_id] = sqlite3.connect(
                    DB_FILE,
                    detect_types=sqlite3.PARSE_DECLTYPES,
                    timeout=30,
                    check_same_thread=False
                )
                self.connections[thread_id].execute("PRAGMA journal_mode=WAL")
                self.connections[thread_id].execute("PRAGMA busy_timeout = 30000")
            return self.connections[thread_id]

    def close_connection(self, thread_id=None):
        """Close one thread connection or all connections."""
        with self._lock:
            if thread_id is None:
                for conn in self.connections.values():
                    conn.close()
                self.connections = {}
            elif thread_id in self.connections:
                self.connections[thread_id].close()
                del self.connections[thread_id]

    def execute_query(self, query, params=None, thread_id=None):
        """Execute a single query and commit."""
        conn = self.get_connection(thread_id)
        cursor = conn.cursor()
        try:
            if params:
                cursor.execute(query, params)
            else:
                cursor.execute(query)
            conn.commit()
            return cursor
        except Exception as e:
            conn.rollback()
            raise e

    def executemany_query(self, query, params_list, thread_id=None):
        """Execute many statements and commit."""
        conn = self.get_connection(thread_id)
        cursor = conn.cursor()
        try:
            cursor.executemany(query, params_list)
            conn.commit()
            return cursor
        except Exception as e:
            conn.rollback()
            raise e


def _new_api_headers() -> Dict[str, str]:
    return {
        "Accept": "application/json, text/plain, */*",
        "Origin": "https://malody.mugzone.net",
        "Referer": "https://malody.mugzone.net/",
        "User-Agent": HEADERS["User-Agent"],
    }


def _new_api_ensure_guest_auth(session: requests.Session, force_refresh: bool = False) -> Dict[str, Any]:
    global _api_auth_cache
    with _api_auth_lock:
        if (
            not force_refresh
            and _api_auth_cache
            and isinstance(_api_auth_cache.get("ts"), float)
            and (time.time() - _api_auth_cache["ts"] < 1800)
        ):
            return _api_auth_cache

        resp = session.get(
            f"{NEW_API_BASE_URL}/web/auth/guest/wt",
            headers=_new_api_headers(),
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        if data.get("code") != 0:
            raise RuntimeError(f"guest auth failed: code={data.get('code')}")

        key = data.get("key") or data.get("token")
        if not key:
            raise RuntimeError("guest auth failed: missing key/token")

        _api_auth_cache = {
            "uid": int(data.get("uid", 1)),
            "key": key,
            "store_key": data.get("storeKey") or data.get("tokenStore") or key,
            "ts": time.time(),
        }
        return _api_auth_cache


def _new_api_get(
    session: requests.Session,
    path: str,
    params: Optional[Dict[str, Any]] = None,
    use_store_key: bool = False,
    retry_on_auth: bool = True,
) -> Dict[str, Any]:
    auth = _new_api_ensure_guest_auth(session)
    request_params = dict(params or {})
    request_params.setdefault("uid", auth["uid"])
    request_params.setdefault("key", auth["store_key"] if use_store_key else auth["key"])

    url = NEW_API_BASE_URL + (path if path.startswith("/") else f"/{path}")
    resp = session.get(url, params=request_params, headers=_new_api_headers(), timeout=30)
    resp.raise_for_status()
    data = resp.json()

    if data.get("code") == -1000 and retry_on_auth:
        _new_api_ensure_guest_auth(session, force_refresh=True)
        return _new_api_get(
            session,
            path=path,
            params=params,
            use_store_key=use_store_key,
            retry_on_auth=False,
        )

    return data


def _safe_int(value: Any, default: int = 0) -> int:
    try:
        if value is None or value == "":
            return default
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _safe_float(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _acquire_mm_run_lock() -> bool:
    os.makedirs(os.path.dirname(MM_RUN_LOCK_FILE), exist_ok=True)
    try:
        fd = os.open(MM_RUN_LOCK_FILE, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(f"{os.getpid()}\n{datetime.now().isoformat()}\n")
        return True
    except FileExistsError:
        return False


def _release_mm_run_lock():
    try:
        if os.path.exists(MM_RUN_LOCK_FILE):
            os.remove(MM_RUN_LOCK_FILE)
    except OSError:
        logger.warning("failed to remove mm run lock: %s", MM_RUN_LOCK_FILE)


def ensure_mm_schema(cursor: sqlite3.Cursor):
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS player_rankings_mm (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            player_id INTEGER,
            uid TEXT NOT NULL,
            mode INTEGER NOT NULL,
            rank INTEGER NOT NULL,
            name TEXT NOT NULL,
            mm_value INTEGER NOT NULL,
            lv INTEGER,
            acc REAL,
            combo INTEGER,
            pc INTEGER,
            crawl_time TIMESTAMP NOT NULL,
            source TEXT DEFAULT 'mm_global',
            FOREIGN KEY (player_id) REFERENCES player_identity (player_id)
        )
        """
    )
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_player_rankings_mm_uid ON player_rankings_mm(uid)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_player_rankings_mm_mode ON player_rankings_mm(mode)")
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_player_rankings_mm_player_mode ON player_rankings_mm(player_id, mode)"
    )

    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS player_mmr_samples (
            sample_id INTEGER PRIMARY KEY AUTOINCREMENT,
            player_id INTEGER,
            uid TEXT NOT NULL,
            mode INTEGER NOT NULL,
            mmr INTEGER NOT NULL,
            mm_rank INTEGER,
            name TEXT,
            crawl_time TIMESTAMP NOT NULL,
            source TEXT NOT NULL,
            FOREIGN KEY (player_id) REFERENCES player_identity (player_id)
        )
        """
    )
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_player_mmr_samples_uid ON player_mmr_samples(uid)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_player_mmr_samples_mode ON player_mmr_samples(mode)")
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_player_mmr_samples_uid_mode_time ON player_mmr_samples(uid, mode, crawl_time)"
    )

    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS player_mmr_daily (
            uid TEXT NOT NULL,
            mode INTEGER NOT NULL,
            day TEXT NOT NULL,
            mmr INTEGER NOT NULL,
            mm_rank INTEGER,
            name TEXT,
            sample_time TIMESTAMP NOT NULL,
            source TEXT NOT NULL,
            PRIMARY KEY (uid, mode, day)
        )
        """
    )
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_player_mmr_daily_day ON player_mmr_daily(day)")

    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS mm_crawl_status (
            task TEXT PRIMARY KEY,
            last_crawled TIMESTAMP,
            last_success TIMESTAMP,
            crawl_count INTEGER DEFAULT 0,
            success_count INTEGER DEFAULT 0,
            last_error TEXT,
            state_json TEXT
        )
        """
    )


def migrate_database():
    """Migrate database: add uid columns for tables if missing."""
    db_manager = DatabaseManager()
    cursor = db_manager.get_connection().cursor()
    logger.info("开始数据库迁移...")
    try:
        # Check the player_identity table.
        cursor.execute("PRAGMA table_info(player_identity)")
        columns = [column[1] for column in cursor.fetchall()]
        if 'uid' not in columns:
            cursor.execute('ALTER TABLE player_identity ADD COLUMN uid TEXT')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_player_identity_uid ON player_identity(uid)')
            logger.info("已添加 uid 字段到 player_identity 表")

        cursor.execute("PRAGMA table_info(player_aliases)")
        columns = [column[1] for column in cursor.fetchall()]
        if 'uid' not in columns:
            cursor.execute('ALTER TABLE player_aliases ADD COLUMN uid TEXT')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_player_aliases_uid ON player_aliases(uid)')
            logger.info("已添加 uid 字段到 player_aliases 表")

        cursor.execute("PRAGMA table_info(player_rankings)")
        columns = [column[1] for column in cursor.fetchall()]
        if 'uid' not in columns:
            cursor.execute('ALTER TABLE player_rankings ADD COLUMN uid TEXT')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_player_rankings_uid ON player_rankings(uid)')
            logger.info("已添加 uid 字段到 player_rankings 表")

        ensure_mm_schema(cursor)
        logger.info("MM/MMR schema ensured")

        db_manager.get_connection().commit()
        # Record migration changes into a markdown file.
        with open('sql_changes.md', 'w', encoding='utf-8') as f:
            f.write("# SQL数据库结构变更记录\n\n")
            f.write("## 版本 2.0 - 添加UID支持\n\n")
            f.write("### 变更内容\n\n")
            f.write("1. 在 `player_identity` 表中添加 `uid` 字段\n")
            f.write("2. 在 `player_aliases` 表中添加 `uid` 字段\n")
            f.write("3. 在 `player_rankings` 表中添加 `uid` 字段\n")
            f.write("4. 为各表的 `uid` 字段创建索引\n\n")
            f.write("### SQL语句\n\n")
            f.write("```sql\n")
            f.write("-- 添加uid字段\n")
            f.write("ALTER TABLE player_identity ADD COLUMN uid TEXT;\n")
            f.write("ALTER TABLE player_aliases ADD COLUMN uid TEXT;\n")
            f.write("ALTER TABLE player_rankings ADD COLUMN uid TEXT;\n\n")
            f.write("-- 创建索引\n")
            f.write("CREATE INDEX idx_player_identity_uid ON player_identity(uid);\n")
            f.write("CREATE INDEX idx_player_aliases_uid ON player_aliases(uid);\n")
            f.write("CREATE INDEX idx_player_rankings_uid ON player_rankings(uid);\n")
            f.write("```\n")
        logger.info("数据库迁移完成，变更已记录到 sql_changes.md")
        return True
    except Exception as e:
        logger.error("数据库迁移失败: %s", e)
        db_manager.get_connection().rollback()
        return False


def init_database():
    """Initialize database and create required table schemas."""
    db_manager = DatabaseManager()
    cursor = db_manager.get_connection().cursor()
    try:
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS player_identity (
            player_id INTEGER PRIMARY KEY AUTOINCREMENT,
            uid TEXT,
            current_name TEXT NOT NULL,
            first_seen TIMESTAMP NOT NULL,
            last_seen TIMESTAMP NOT NULL
        )
        ''')

        cursor.execute('''
        CREATE TABLE IF NOT EXISTS player_aliases (
            alias_id INTEGER PRIMARY KEY AUTOINCREMENT,
            player_id INTEGER NOT NULL,
            uid TEXT,
            alias TEXT NOT NULL,
            first_seen TIMESTAMP NOT NULL,
            last_seen TIMESTAMP NOT NULL,
            FOREIGN KEY (player_id) REFERENCES player_identity (player_id)
        )
        ''')

        cursor.execute('''
        CREATE TABLE IF NOT EXISTS player_rankings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            player_id INTEGER,
            uid TEXT,
            mode INTEGER NOT NULL,
            rank INTEGER NOT NULL,
            name TEXT NOT NULL,
            lv INTEGER,
            exp INTEGER,
            acc REAL,
            combo INTEGER,
            pc INTEGER,
            crawl_time TIMESTAMP NOT NULL,
            FOREIGN KEY (player_id) REFERENCES player_identity (player_id)
        )
        ''')

        cursor.execute('''
        CREATE TABLE IF NOT EXISTS import_metadata (
            mode INTEGER PRIMARY KEY,
            last_import_time TIMESTAMP
        )
        ''')

        cursor.execute('''
        CREATE TABLE IF NOT EXISTS player_config (
            config_id INTEGER PRIMARY KEY AUTOINCREMENT,
            player_identifier TEXT NOT NULL UNIQUE,
            player_name TEXT,
            is_active BOOLEAN DEFAULT 1,
            priority INTEGER DEFAULT 5,
            notes TEXT,
            created_time TIMESTAMP NOT NULL,
            last_updated TIMESTAMP NOT NULL
        )
        ''')

        cursor.execute('''
        CREATE TABLE IF NOT EXISTS player_crawl_status (
            player_identifier TEXT PRIMARY KEY,
            last_crawled TIMESTAMP,
            crawl_count INTEGER DEFAULT 0,
            success_count INTEGER DEFAULT 0,
            last_error TEXT
        )
        ''')

        cursor.execute('CREATE INDEX IF NOT EXISTS idx_player_identity_uid ON player_identity(uid)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_player_aliases_uid ON player_aliases(uid)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_player_rankings_uid ON player_rankings(uid)')
        ensure_mm_schema(cursor)

        # Initialize import metadata rows.
        for mode in MODES:
            cursor.execute(
                "INSERT OR IGNORE INTO import_metadata (mode, last_import_time) VALUES (?, NULL)",
                (mode,)
            )

        db_manager.get_connection().commit()
        logger.info("数据库初始化完成")
        migrate_database()  # 尝试添加 uid 字段（若已存在则跳过）
    except Exception as e:
        logger.error("数据库初始化失败: %s", e)
        raise


def resolve_player_identity(name, crawl_time, uid=None):
    """
    解析玩家身份，优先使用 uid，处理改名情况。
    返回 player_id。
    """
    db_manager = DatabaseManager()
    cursor = db_manager.get_connection().cursor()
    try:
        player_id = None
        # Prefer uid-based lookup first.
        if uid:
            try:
                cursor.execute(
                    "SELECT player_id FROM player_identity WHERE uid = ?",
                    (uid,)
                )
                result = cursor.fetchone()
                if result:
                    player_id = result[0]
                    # Refresh last-seen timestamp and current display name.
                    cursor.execute(
                        "UPDATE player_identity SET last_seen = ?, current_name = ? WHERE player_id = ?",
                        (crawl_time, name, player_id)
                    )
                    # Insert alias record when missing.
                    cursor.execute(
                        "SELECT alias_id FROM player_aliases WHERE player_id = ? AND alias = ?",
                        (player_id, name)
                    )
                    if not cursor.fetchone():
                        cursor.execute(
                            "INSERT INTO player_aliases (player_id, uid, alias, first_seen, last_seen) VALUES (?, ?, ?, ?, ?)",
                            (player_id, uid, name, crawl_time, crawl_time)
                        )
                    else:
                        cursor.execute(
                            "UPDATE player_aliases SET last_seen = ? WHERE player_id = ? AND alias = ?",
                            (crawl_time, player_id, name)
                        )
            except sqlite3.OperationalError as e:
                if "no such column: uid" in str(e):
                    logger.warning("uid 列不存在，回退到名称查找")
                else:
                    raise

        # If uid lookup misses, fall back to alias lookup.
        if not player_id:
            cursor.execute(
                "SELECT player_id FROM player_aliases WHERE alias = ?",
                (name,)
            )
            result = cursor.fetchone()
            if result:
                player_id = result[0]
                # If uid is available, backfill uid into identity/alias tables.
                if uid:
                    try:
                        cursor.execute(
                            "UPDATE player_identity SET uid = ? WHERE player_id = ?",
                            (uid, player_id)
                        )
                        cursor.execute(
                            "UPDATE player_aliases SET uid = ? WHERE player_id = ?",
                            (uid, player_id)
                        )
                    except sqlite3.OperationalError as e:
                        if "no such column: uid" in str(e):
                            logger.warning("uid 列不存在，跳过 uid 更新")
                        else:
                            raise
                cursor.execute(
                    "UPDATE player_aliases SET last_seen = ? WHERE alias = ?",
                    (crawl_time, name)
                )
                cursor.execute(
                    "UPDATE player_identity SET last_seen = ?, current_name = ? WHERE player_id = ?",
                    (crawl_time, name, player_id)
                )
            else:
                # Create a brand-new player identity.
                cursor.execute(
                    "INSERT INTO player_identity (uid, current_name, first_seen, last_seen) VALUES (?, ?, ?, ?)",
                    (uid, name, crawl_time, crawl_time)
                )
                player_id = cursor.lastrowid
                cursor.execute(
                    "INSERT INTO player_aliases (player_id, uid, alias, first_seen, last_seen) VALUES (?, ?, ?, ?, ?)",
                    (player_id, uid, name, crawl_time, crawl_time)
                )
        db_manager.get_connection().commit()
        return player_id
    except Exception as e:
        logger.error("解析玩家身份失败: %s", e)
        db_manager.get_connection().rollback()
        return None


def link_player_aliases(original_name, new_name, change_time):
    """Manually link two names for one player (rename handling)."""
    db_manager = DatabaseManager()
    cursor = db_manager.get_connection().cursor()
    try:
        cursor.execute(
            "SELECT player_id FROM player_aliases WHERE alias = ?",
            (original_name,)
        )
        result = cursor.fetchone()
        if not result:
            logger.error("找不到原始名字: %s", original_name)
            return False
        player_id = result[0]

        cursor.execute(
            "SELECT player_id FROM player_aliases WHERE alias = ?",
            (new_name,)
        )
        result = cursor.fetchone()
        if result:
            # New name belongs to another player_id; merge required.
            old_player_id = result[0]
            cursor.execute(
                "UPDATE player_rankings SET player_id = ? WHERE player_id = ?",
                (player_id, old_player_id)
            )
            cursor.execute(
                "UPDATE player_aliases SET player_id = ? WHERE player_id = ?",
                (player_id, old_player_id)
            )
            cursor.execute(
                "DELETE FROM player_identity WHERE player_id = ?",
                (old_player_id,)
            )
        else:
            # New name has not appeared before; add alias directly.
            cursor.execute(
                "INSERT INTO player_aliases (player_id, alias, first_seen, last_seen) VALUES (?, ?, ?, ?)",
                (player_id, new_name, change_time, change_time)
            )
        cursor.execute(
            "UPDATE player_identity SET current_name = ? WHERE player_id = ?",
            (new_name, player_id)
        )
        db_manager.get_connection().commit()
        logger.info("成功关联玩家改名: %s -> %s", original_name, new_name)
        return True
    except Exception as e:
        logger.error("处理玩家改名失败: %s", e)
        db_manager.get_connection().rollback()
        return False


def parse_player_list(html):
    """Parse leaderboard page and return player rows (with player ID)."""
    soup = BeautifulSoup(html, "html.parser")
    players = []

    # Handle top 3 entries (item-top).
    top_items = soup.select("div.item-top")
    for item in top_items:
        label_tag = item.select_one("i.label")
        rank = None
        if label_tag and label_tag.has_attr("class"):
            for c in label_tag["class"]:
                if c.startswith("top-"):
                    rank = c.replace("top-", "")
                    break

        name_tag = item.select_one("span.name a")
        lv_tag = item.select_one("span.lv")
        acc_tag = item.select_one("span.acc")
        combo_tag = item.select_one("span.combo")
        pc_tag = item.select_one("span.pc, span[class*=pc]")

        player_id = None
        if name_tag and name_tag.has_attr('href'):
            href = name_tag['href']
            match = re.search(r'/accounts/user/(\d+)', href)
            if match:
                player_id = match.group(1)

        level = None
        exp = None
        if lv_tag:
            lv_text = lv_tag.text.strip()
            if '-' in lv_text:
                parts = lv_text.split('-')
                level = parts[0].replace("Lv.", "").strip()
                exp = parts[1].strip()
            else:
                level = lv_text.replace("Lv.", "").strip()

        acc_text = None
        if acc_tag:
            acc_text = acc_tag.text.replace("Acc:", "").replace("%", "").strip()

        combo_text = None
        if combo_tag:
            combo_text = combo_tag.text.replace("Combo:", "").strip()

        playcount = None
        if pc_tag:
            # Keep this parser prefix-agnostic to avoid locale/encoding regressions.
            pc_text = pc_tag.text.strip()
            digits = ''.join(filter(str.isdigit, pc_text))
            if digits:
                playcount = int(digits)

        players.append({
            "rank": rank,
            "name": name_tag.text.strip() if name_tag else None,
            "player_id": player_id,
            "lv": level,
            "exp": exp,
            "acc": acc_text,
            "combo": combo_text,
            "pc": playcount
        })

    # Handle rank 4+ entries (div.item).
    list_items = soup.select("div.item")
    for item in list_items:
        rank_tag = item.select_one("span.rank")
        rank = rank_tag.text.strip() if rank_tag else None

        name_tag = item.select_one("span.name a")
        lv_tag = item.select_one("span.lv")
        exp_tag = item.select_one("span.exp")
        acc_tag = item.select_one("span.acc")
        pc_tag = item.select_one("span.pc, span[class*=pc]")
        combo_tag = item.select_one("span.combo")

        player_id = None
        if name_tag and name_tag.has_attr('href'):
            href = name_tag['href']
            match = re.search(r'/accounts/user/(\d+)', href)
            if match:
                player_id = match.group(1)

        acc_text = None
        if acc_tag:
            acc_text = acc_tag.text.replace("%", "").strip()

        playcount = None
        if pc_tag:
            pc_text = pc_tag.text.strip()
            digits = ''.join(filter(str.isdigit, pc_text))
            if digits:
                playcount = int(digits)

        players.append({
            "rank": rank,
            "name": name_tag.text.strip() if name_tag else None,
            "player_id": player_id,
            "lv": lv_tag.text.strip() if lv_tag else None,
            "exp": exp_tag.text.strip() if exp_tag else None,
            "acc": acc_text,
            "combo": combo_tag.text.strip() if combo_tag else None,
            "pc": playcount
        })

    # Type conversion and data cleanup.
    processed_players = []
    for p in players:
        try:
            rank = int(p["rank"]) if p["rank"] else None
        except:
            rank = None
        try:
            lv = int(p["lv"]) if p["lv"] else 0
        except:
            lv = 0
        try:
            exp = int(p["exp"]) if p["exp"] else 0
        except:
            exp = 0
        try:
            acc = float(p["acc"]) if p["acc"] else 0.0
        except:
            acc = 0.0
        try:
            combo = int(p["combo"]) if p["combo"] else 0
        except:
            combo = 0
        pc = p["pc"] if p["pc"] is not None else 0

        if rank is not None:
            processed_players.append({
                "rank": rank,
                "name": p["name"],
                "player_id": p["player_id"],
                "lv": lv,
                "exp": exp,
                "acc": acc,
                "combo": combo,
                "pc": pc
            })
    return processed_players


def parse_player_profile(html, player_id):
    """Parse player profile page and return rankings for all modes."""
    soup = BeautifulSoup(html, "html.parser")
    player_data = []

    name_tag = soup.select_one("div.user_head .name span")
    player_name = name_tag.text.strip() if name_tag else f"玩家_{player_id}"

    rank_items = soup.select("div.rank .item")
    for item in rank_items:
        try:
            img_tag = item.select_one("img")
            if not img_tag or not img_tag.has_attr('src'):
                continue
            src = img_tag['src']
            mode = None
            if 'mode-0.png' in src:
                mode = 0
            elif 'mode-1.png' in src:
                mode = 1
            elif 'mode-2.png' in src:
                mode = 2
            elif 'mode-3.png' in src:
                mode = 3
            elif 'mode-4.png' in src:
                mode = 4
            elif 'mode-5.png' in src:
                mode = 5
            elif 'mode-6.png' in src:
                mode = 6
            elif 'mode-7.png' in src:
                mode = 7
            elif 'mode-8.png' in src:
                mode = 8
            elif 'mode-9.png' in src:
                mode = 9
            else:
                continue

            rank_tag = item.select_one("p.rank")
            if not rank_tag:
                continue
            rank_text = rank_tag.text.strip()
            if rank_text.startswith('#'):
                try:
                    rank = int(rank_text[1:].replace(',', ''))
                except:
                    continue
            else:
                continue

            data_spans = item.select("p span")
            exp = 0
            playcount = 0
            acc = 0.0
            combo = 0
            for span in data_spans:
                text = span.text.strip()
                if text.startswith('Exp.'):
                    try:
                        exp = int(text.replace('Exp.', '').strip().replace(',', ''))
                    except:
                        pass
                elif text.startswith('Playcount:'):
                    try:
                        playcount = int(text.replace('Playcount:', '').strip())
                    except:
                        pass
                elif text.startswith('Acc.'):
                    try:
                        acc = float(text.replace('Acc.', '').replace('%', '').strip())
                    except:
                        pass
                elif text.startswith('Combo:'):
                    try:
                        combo = int(text.replace('Combo:', '').strip())
                    except:
                        pass

            player_data.append({
                "rank": rank,
                "name": player_name,
                "lv": 0,
                "exp": exp,
                "acc": acc,
                "combo": combo,
                "pc": playcount,
                "mode": mode
            })
        except Exception as e:
            logger.warning("解析玩家 %s 模式 %s 数据时出错: %s", player_id, mode, e)
            continue
    return player_data


def crawl_player_profile(session, player_identifier):
    """Fetch one player profile page and parse ranking records."""
    try:
        if not player_identifier.isdigit():
            logger.warning("玩家标识符必须是数字ID: %s", player_identifier)
            return None
        url = PLAYER_PROFILE_URL.format(player_id=player_identifier)
        resp = session.get(url, timeout=30)
        resp.raise_for_status()
        player_data = parse_player_profile(resp.text, player_identifier)
        return player_data
    except requests.exceptions.RequestException as e:
        logger.error("爬取玩家 %s 个人主页失败: %s", player_identifier, e)
        return None
    except Exception as e:
        logger.error("处理玩家 %s 数据时出错: %s", player_identifier, e)
        return None


def get_excel_filename(mode):
    """Return the Excel filename for the given mode."""
    if mode == 0:
        return "key.xlsx"
    elif mode == 3:
        return "catch.xlsx"
    else:
        return f"mode{mode}.xlsx"


def crawl_mode_player(session, mode):
    """Fetch one mode leaderboard and return a DataFrame."""
    url = BASE_URL.format(mode=mode)
    try:
        resp = session.get(url, timeout=30)
        resp.raise_for_status()
    except requests.exceptions.RequestException as e:
        logger.error("模式 %d 请求失败: %s", mode, e)
        notify_user("排行榜爬取异常", f"模式 {mode} 请求失败: {e}", "error")
        return pd.DataFrame()

    players = parse_player_list(resp.text)
    df = pd.DataFrame(players)
    if not df.empty:
        df = df[df['rank'].notnull()]
        df['rank'] = df['rank'].astype(int)
        df = df.sort_values('rank').reset_index(drop=True)
    return df


def save_data_to_excel(mode, df, timestamp):
    """Save DataFrame to Excel (only when Excel output is enabled)."""
    if df.empty:
        logger.warning("模式 %d 无有效数据，跳过保存", mode)
        return

    filename = get_excel_filename(mode)
    sheet_name = f"mode_{mode}"

    try:
        if not os.path.exists(filename):
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                pd.DataFrame().to_excel(writer, sheet_name=sheet_name)

        wb = load_workbook(filename)
        if sheet_name not in wb.sheetnames:
            with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
                pd.DataFrame().to_excel(writer, sheet_name=sheet_name)

        sub_sheets = [s for s in wb.sheetnames if s.startswith(f"{sheet_name}_")] if wb.sheetnames else []
        latest_sheet = None
        latest_time = None
        for s in sub_sheets:
            try:
                dt_str = s.replace(f"{sheet_name}_", "")
                dt = datetime.strptime(dt_str, "%Y-%m-%d_%H-%M")
                if latest_time is None or dt > latest_time:
                    latest_time = dt
                    latest_sheet = s
            except:
                continue

        if latest_sheet:
            df_prev = pd.read_excel(filename, sheet_name=latest_sheet)
            if not df_prev.empty and df_prev.equals(df):
                logger.info("模式 %d 数据未变化，跳过保存", mode)
                return

        sub_sheet_name = f"{sheet_name}_{timestamp.strftime('%Y-%m-%d_%H-%M')}"
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, sheet_name=sub_sheet_name, index=False)
        logger.info("模式 %d 数据保存到 %s -> %s", mode, filename, sub_sheet_name)
    except Exception as e:
        logger.exception("保存模式 %d 数据到Excel失败", mode)


def save_player_ranking_record(player_id, uid, mode, rank, name, lv, exp, acc, combo, pc, crawl_time, source):
    """
    区间模型排行榜保存（统一处理所有来源）。
    返回值：
        'new'          : 首次插入
        'diff_insert'  : 状态变化，插入新 start
        'update'       : 延长 end
        'same_insert'  : 插入 end
        'update_fill'  : 填充 lv
    """
    db_manager = DatabaseManager()
    conn = db_manager.get_connection()
    cursor = conn.cursor()

    current_core = (rank, name, exp, acc, combo, pc)

    # Load only the latest two rows for interval-state comparison.
    cursor.execute('''
        SELECT id, rank, name, lv, exp, acc, combo, pc, crawl_time
        FROM player_rankings
        WHERE player_id = ? AND mode = ?
        ORDER BY crawl_time DESC
        LIMIT 2
    ''', (player_id, mode))

    rows = cursor.fetchall()

    # No history yet -> insert the first start record.
    if not rows:
        cursor.execute('''
            INSERT INTO player_rankings
            (player_id, uid, mode, rank, name, lv, exp, acc, combo, pc, crawl_time)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (player_id, uid, mode, rank, name, lv, exp, acc, combo, pc, crawl_time))
        conn.commit()
        return 'new'

    # Latest stored record.
    last = rows[0]
    last_id, last_rank, last_name, last_lv, last_exp, last_acc, last_combo, last_pc, last_time = last
    last_core = (last_rank, last_name, last_exp, last_acc, last_combo, last_pc)

    # Determine whether an end marker already exists (latest two rows equal).
    has_two_same = False
    if len(rows) == 2:
        second = rows[1]
        second_core = (second[1], second[2], second[4], second[5], second[6], second[7])
        if second_core == last_core:
            has_two_same = True

    # Current core fields are identical to latest state.
    if last_core == current_core:
        # Guard against legacy/source bug: lv may temporarily bounce to 0 even when
        # other core fields are unchanged. Keep the known non-zero lv in this case.
        if last_lv and last_lv > 0 and (lv is None or lv == 0):
            lv = last_lv

        # Backfill lv only when previous lv is missing (0).
        if last_lv == 0 and lv != 0:
            cursor.execute('UPDATE player_rankings SET lv = ?, crawl_time = ? WHERE id = ?', (lv, crawl_time, last_id))
            conn.commit()
            return 'update_fill'

        if not has_two_same:
            # Insert a new end record for unchanged status.
            cursor.execute('''
                INSERT INTO player_rankings
                (player_id, uid, mode, rank, name, lv, exp, acc, combo, pc, crawl_time)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (player_id, uid, mode, rank, name, lv, exp, acc, combo, pc, crawl_time))
            conn.commit()
            return 'same_insert'
        else:
            # Extend end by updating latest crawl_time only.
            cursor.execute('UPDATE player_rankings SET crawl_time = ? WHERE id = ?', (crawl_time, last_id))
            conn.commit()
            return 'update'

    # State changed -> start a new interval segment.
    cursor.execute('''
        INSERT INTO player_rankings
        (player_id, uid, mode, rank, name, lv, exp, acc, combo, pc, crawl_time)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (player_id, uid, mode, rank, name, lv, exp, acc, combo, pc, crawl_time))

    conn.commit()
    return 'diff_insert'


def save_player_ranking_mm_record(
    player_id: int,
    uid: str,
    mode: int,
    rank: int,
    name: str,
    mm_value: int,
    lv: int,
    acc: float,
    combo: int,
    pc: int,
    crawl_time: datetime,
    source: str,
) -> str:
    """
    MM 排行区间模型保存。
    """
    db_manager = DatabaseManager()
    conn = db_manager.get_connection()
    cursor = conn.cursor()

    current_core = (rank, name, mm_value, lv, acc, combo, pc)
    cursor.execute(
        """
        SELECT id, rank, name, mm_value, lv, acc, combo, pc, crawl_time
        FROM player_rankings_mm
        WHERE player_id = ? AND mode = ?
        ORDER BY crawl_time DESC
        LIMIT 2
        """,
        (player_id, mode),
    )
    rows = cursor.fetchall()

    if not rows:
        cursor.execute(
            """
            INSERT INTO player_rankings_mm
            (player_id, uid, mode, rank, name, mm_value, lv, acc, combo, pc, crawl_time, source)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (player_id, uid, mode, rank, name, mm_value, lv, acc, combo, pc, crawl_time, source),
        )
        conn.commit()
        return "new"

    last = rows[0]
    last_id = last[0]
    last_core = (last[1], last[2], last[3], last[4], last[5], last[6], last[7])
    has_two_same = False
    if len(rows) == 2:
        second = rows[1]
        second_core = (second[1], second[2], second[3], second[4], second[5], second[6], second[7])
        has_two_same = second_core == last_core

    if last_core == current_core:
        if not has_two_same:
            cursor.execute(
                """
                INSERT INTO player_rankings_mm
                (player_id, uid, mode, rank, name, mm_value, lv, acc, combo, pc, crawl_time, source)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (player_id, uid, mode, rank, name, mm_value, lv, acc, combo, pc, crawl_time, source),
            )
            conn.commit()
            return "same_insert"
        cursor.execute("UPDATE player_rankings_mm SET crawl_time = ? WHERE id = ?", (crawl_time, last_id))
        conn.commit()
        return "update"

    cursor.execute(
        """
        INSERT INTO player_rankings_mm
        (player_id, uid, mode, rank, name, mm_value, lv, acc, combo, pc, crawl_time, source)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (player_id, uid, mode, rank, name, mm_value, lv, acc, combo, pc, crawl_time, source),
    )
    conn.commit()
    return "diff_insert"


def save_player_mmr_sample(
    player_id: Optional[int],
    uid: str,
    mode: int,
    mmr: int,
    mm_rank: Optional[int],
    name: Optional[str],
    crawl_time: datetime,
    source: str,
) -> str:
    db_manager = DatabaseManager()
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    day = crawl_time.date().isoformat()

    # Align with ranking dedup model: keep two boundary rows per stable segment.
    cursor.execute(
        """
        SELECT sample_id, mmr, mm_rank, name, source
        FROM player_mmr_samples
        WHERE uid = ? AND mode = ?
        ORDER BY crawl_time DESC
        LIMIT 2
        """,
        (uid, mode),
    )
    rows = cursor.fetchall()
    if not rows:
        cursor.execute(
            """
            INSERT INTO player_mmr_samples
            (player_id, uid, mode, mmr, mm_rank, name, crawl_time, source)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (player_id, uid, mode, mmr, mm_rank, name, crawl_time, source),
        )
        sample_op = "new"
    else:
        last_sample_id, last_mmr, last_mm_rank, last_name, last_source = rows[0]
        has_two_same = False
        if len(rows) == 2:
            second = rows[1]
            has_two_same = (
                int(second[1] or 0) == int(last_mmr or 0)
                and (second[2] if second[2] is not None else None) == (last_mm_rank if last_mm_rank is not None else None)
                and (second[3] or None) == (last_name or None)
                and (second[4] or "") == (last_source or "")
            )

        same_as_last = (
            int(last_mmr or 0) == int(mmr)
            and (last_mm_rank if last_mm_rank is not None else None) == (mm_rank if mm_rank is not None else None)
            and (last_name or None) == (name or None)
            and (last_source or "") == (source or "")
        )
        if same_as_last:
            if not has_two_same:
                cursor.execute(
                    """
                    INSERT INTO player_mmr_samples
                    (player_id, uid, mode, mmr, mm_rank, name, crawl_time, source)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (player_id, uid, mode, mmr, mm_rank, name, crawl_time, source),
                )
                sample_op = "same_insert"
            else:
                cursor.execute(
                    """
                    UPDATE player_mmr_samples
                    SET crawl_time = ?
                    WHERE sample_id = ?
                    """,
                    (crawl_time, last_sample_id),
                )
                sample_op = "update"
        else:
            cursor.execute(
                """
                INSERT INTO player_mmr_samples
                (player_id, uid, mode, mmr, mm_rank, name, crawl_time, source)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (player_id, uid, mode, mmr, mm_rank, name, crawl_time, source),
            )
            sample_op = "diff_insert"

    cursor.execute(
        """
        INSERT INTO player_mmr_daily
        (uid, mode, day, mmr, mm_rank, name, sample_time, source)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(uid, mode, day) DO UPDATE SET
            mmr = excluded.mmr,
            mm_rank = excluded.mm_rank,
            name = excluded.name,
            sample_time = excluded.sample_time,
            source = excluded.source
        WHERE excluded.sample_time >= player_mmr_daily.sample_time
        """,
        (uid, mode, day, mmr, mm_rank, name, crawl_time, source),
    )
    conn.commit()
    return sample_op

def update_mm_crawl_status(
    task: str,
    success: bool,
    error: Optional[str] = None,
    state: Optional[Dict[str, Any]] = None,
):
    db_manager = DatabaseManager()
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    now = datetime.now()
    state_json = json.dumps(state or {}, ensure_ascii=False)
    cursor.execute(
        """
        INSERT OR REPLACE INTO mm_crawl_status
        (task, last_crawled, last_success, crawl_count, success_count, last_error, state_json)
        VALUES (
            ?,
            ?,
            CASE WHEN ? THEN ? ELSE (SELECT last_success FROM mm_crawl_status WHERE task = ?) END,
            COALESCE((SELECT crawl_count FROM mm_crawl_status WHERE task = ?), 0) + 1,
            COALESCE((SELECT success_count FROM mm_crawl_status WHERE task = ?), 0) + CASE WHEN ? THEN 1 ELSE 0 END,
            ?,
            ?
        )
        """,
        (
            task,
            now,
            1 if success else 0,
            now,
            task,
            task,
            task,
            1 if success else 0,
            None if success else error,
            state_json,
        ),
    )
    conn.commit()


def fetch_global_mode(
    session: requests.Session,
    mode: int,
    limit: int = DEFAULT_MM_LIMIT,
    mm: int = 0,
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    from_offset = 0
    visited_offsets: Set[int] = set()
    while len(rows) < max(limit, 0):
        if from_offset in visited_offsets:
            break
        visited_offsets.add(from_offset)

        payload = _new_api_get(
            session,
            "/ranking/global",
            {"mode": mode, "from": from_offset, "mm": 1 if mm else 0},
        )
        if payload.get("code") != 0:
            raise RuntimeError(f"ranking/global failed: code={payload.get('code')}")
        data = payload.get("data") or []
        if not data:
            break
        for item in data:
            rows.append(item)
            if len(rows) >= limit:
                break

        if not payload.get("hasMore"):
            break
        next_offset = payload.get("next")
        if isinstance(next_offset, int):
            from_offset = next_offset
        else:
            from_offset += len(data)

    return rows


def fetch_mm_global_mode(
    session: requests.Session,
    mode: int,
    limit: int = DEFAULT_MM_LIMIT,
) -> List[Dict[str, Any]]:
    return fetch_global_mode(session=session, mode=mode, limit=limit, mm=1)


def crawl_mode_player_newapi(
    session: requests.Session,
    mode: int,
    limit: int = DEFAULT_MM_LIMIT,
) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:
    try:
        exp_rows = fetch_global_mode(session=session, mode=mode, limit=limit, mm=0)
    except Exception as e:
        logger.error("newapi exp ranking crawl failed for mode %d: %s", mode, e)
        notify_user("排行榜API爬取失败", f"模式 {mode} newapi请求失败: {e}", "error")
        return pd.DataFrame(), []

    players: List[Dict[str, Any]] = []
    for item in exp_rows:
        uid_int = _safe_int(item.get("uid"), 0)
        if uid_int <= 0:
            continue
        players.append(
            {
                "rank": _safe_int(item.get("rank"), 0),
                "name": item.get("username") or f"uid_{uid_int}",
                "player_id": str(uid_int),
                "lv": _safe_int(item.get("level"), 0),
                "exp": _safe_int(item.get("value"), 0),
                "acc": _safe_float(item.get("acc"), 0.0),
                "combo": _safe_int(item.get("combo"), 0),
                "pc": _safe_int(item.get("playcount"), 0),
            }
        )

    df = pd.DataFrame(players)
    if not df.empty:
        df = df[df["rank"] > 0]
        if df["rank"].duplicated().any():
            dup_count = int(df["rank"].duplicated().sum())
            logger.warning(
                "newapi mode %d contains duplicated ranks (%d), keep first row per rank.",
                mode,
                dup_count,
            )
            df = df.drop_duplicates(subset=["rank"], keep="first")
        df = df.sort_values("rank").reset_index(drop=True)

    mm_rows: List[Dict[str, Any]] = []
    try:
        mm_rows = fetch_mm_global_mode(session=session, mode=mode, limit=limit)
    except Exception as e:
        logger.warning("newapi mm ranking crawl failed for mode %d: %s", mode, e)
        mm_rows = []

    return df, mm_rows


def save_mm_ranking_rows(
    mode: int,
    rows: List[Dict[str, Any]],
    crawl_time: datetime,
    source: str = "mm_global",
) -> Dict[str, int]:
    row_stats = {"new": 0, "same_insert": 0, "update": 0, "diff_insert": 0}
    seen_ranks: Set[int] = set()
    seen_uids: Set[str] = set()
    normalized_rows: List[Dict[str, Any]] = []
    dropped_dup_rank = 0
    dropped_dup_uid = 0
    dropped_invalid = 0

    for item in rows:
        uid = str(_safe_int(item.get("uid"), 0))
        rank = _safe_int(item.get("rank"), 0)
        if uid == "0" or rank <= 0:
            dropped_invalid += 1
            continue
        if rank in seen_ranks:
            dropped_dup_rank += 1
            continue
        if uid in seen_uids:
            dropped_dup_uid += 1
            continue
        seen_ranks.add(rank)
        seen_uids.add(uid)
        normalized_rows.append(item)

    saved_rows = 0
    for item in normalized_rows:
        uid = str(_safe_int(item.get("uid"), 0))
        if uid == "0":
            continue
        name = item.get("username") or f"uid_{uid}"
        player_id = resolve_player_identity(name, crawl_time, uid=uid)
        if player_id is None:
            continue
        op = save_player_ranking_mm_record(
            player_id=player_id,
            uid=uid,
            mode=mode,
            rank=_safe_int(item.get("rank"), 0),
            name=name,
            mm_value=_safe_int(item.get("value"), 0),
            lv=_safe_int(item.get("level"), 0),
            acc=_safe_float(item.get("acc"), 0.0),
            combo=_safe_int(item.get("combo"), 0),
            pc=_safe_int(item.get("playcount"), 0),
            crawl_time=crawl_time,
            source=source,
        )
        if op in row_stats:
            row_stats[op] += 1
        saved_rows += 1
    return {
        "rows": len(rows),
        "saved_rows": saved_rows,
        "dropped_dup_rank": dropped_dup_rank,
        "dropped_dup_uid": dropped_dup_uid,
        "dropped_invalid": dropped_invalid,
        **row_stats,
    }


def fetch_player_mmr_from_api(session: requests.Session, uid: str) -> List[Dict[str, Any]]:
    payload = _new_api_get(session, "/ranking/player/all", {"touid": _safe_int(uid, -1)})
    if payload.get("code") != 0:
        raise RuntimeError(f"ranking/player/all failed: code={payload.get('code')}")
    data = payload.get("data") or []
    rows: List[Dict[str, Any]] = []
    for item in data:
        mode = _safe_int(item.get("mode"), -1)
        if mode < 0:
            continue
        mmr = _safe_int(item.get("grade"), -1)
        if mmr < 0:
            continue
        rows.append(
            {
                "uid": str(uid),
                "mode": mode,
                "mmr": mmr,
                "mm_rank": _safe_int(item.get("gradeRank"), 0) or None,
                "rank": _safe_int(item.get("rank"), 0) or None,
                "name": None,
            }
        )
    return rows


def fetch_player_mmr_from_page(session: requests.Session, uid: str) -> List[Dict[str, Any]]:
    url = PLAYER_PROFILE_URL.format(player_id=uid)
    response = session.get(url, timeout=30)
    if response.status_code != 200:
        return []
    soup = BeautifulSoup(response.text, "html.parser")
    rows: List[Dict[str, Any]] = []
    for item in soup.select("div.rank .item"):
        img = item.select_one("img")
        if not img or not img.has_attr("src"):
            continue
        mode_match = re.search(r"mode-(\d+)\.png", img["src"])
        if not mode_match:
            continue
        mode = _safe_int(mode_match.group(1), -1)
        if mode < 0:
            continue
        text = item.get_text(" ", strip=True)
        mmr_match = re.search(r"(?:MMR|MM|Grade)[\s:：]*([0-9,]+)", text, re.IGNORECASE)
        rank_match = re.search(r"(?:MMRank|GradeRank)[\s:：]*#?([0-9,]+)", text, re.IGNORECASE)
        if not mmr_match:
            continue
        rows.append(
            {
                "uid": str(uid),
                "mode": mode,
                "mmr": _safe_int(mmr_match.group(1).replace(",", ""), -1),
                "mm_rank": _safe_int(rank_match.group(1).replace(",", ""), 0) if rank_match else None,
                "name": None,
            }
        )
    return [row for row in rows if row["mmr"] >= 0]


def get_recent_mm_tracked_uids(limit_per_mode: int = DEFAULT_MM_LIMIT) -> Set[str]:
    db_manager = DatabaseManager()
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    tracked: Set[str] = set()
    for mode in MODES:
        cursor.execute(
            """
            SELECT uid
            FROM player_rankings_mm
            WHERE mode = ?
              AND crawl_time = (SELECT MAX(crawl_time) FROM player_rankings_mm WHERE mode = ?)
              AND rank <= ?
            """,
            (mode, mode, limit_per_mode),
        )
        tracked.update(str(row[0]) for row in cursor.fetchall() if row and row[0])
    return tracked


def get_recent_exp_tracked_uids(limit_per_mode: int = DEFAULT_MM_LIMIT) -> Set[str]:
    """
    Extract top-N UIDs from latest EXP ranking per mode to expand MMR fetch coverage for non-MM players."""
    db_manager = DatabaseManager()
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    tracked: Set[str] = set()
    for mode in MODES:
        cursor.execute(
            """
            SELECT uid
            FROM player_rankings
            WHERE mode = ?
              AND crawl_time = (SELECT MAX(crawl_time) FROM player_rankings WHERE mode = ?)
              AND rank <= ?
              AND uid IS NOT NULL
              AND uid != ''
            """,
            (mode, mode, limit_per_mode),
        )
        tracked.update(str(row[0]) for row in cursor.fetchall() if row and row[0])
    return tracked


def get_player_id_by_uid(uid: str) -> Optional[int]:
    db_manager = DatabaseManager()
    cursor = db_manager.get_connection().cursor()
    cursor.execute("SELECT player_id FROM player_identity WHERE uid = ?", (uid,))
    row = cursor.fetchone()
    return row[0] if row else None


def run_mm_sync_cycle(
    mm_limit: int = DEFAULT_MM_LIMIT,
    include_mm_ranking: bool = True,
    include_mmr: bool = True,
    include_exp_ranking_uids: bool = True,
    exp_limit: int = DEFAULT_MM_LIMIT,
    seed_uids: Optional[Set[str]] = None,
    mmr_min_interval_minutes: int = 20,
) -> Dict[str, Any]:
    if not _acquire_mm_run_lock():
        logger.warning("MM sync is already running, skip this cycle.")
        return {"skipped": True, "reason": "locked"}

    session = requests.Session()
    session.cookies.update(COOKIES)
    session.headers.update(HEADERS)
    session.mount("https://", requests.adapters.HTTPAdapter(max_retries=3))

    stats: Dict[str, Any] = {
        "mm_rows": 0,
        "mm_modes_ok": 0,
        "mm_modes_fail": 0,
        "mmr_users_ok": 0,
        "mmr_users_fail": 0,
        "mmr_samples": 0,
        "mmr_sample_new": 0,
        "mmr_sample_diff_insert": 0,
        "mmr_sample_same_insert": 0,
        "mmr_sample_update": 0,
        "mmr_uid_pool": 0,
        "mmr_uid_from_mm_top": 0,
        "mmr_uid_from_exp_top": 0,
        "mmr_uid_from_manual": 0,
        "mmr_uid_from_seed": 0,
        "mmr_users_skipped_recent": 0,
    }
    mm_uids: Set[str] = set()
    crawl_time = datetime.now()
    logger.info(
        "MM sync start: mm_limit=%d include_mm_ranking=%s include_mmr=%s mmr_min_interval_minutes=%d",
        mm_limit,
        include_mm_ranking,
        include_mmr,
        mmr_min_interval_minutes,
    )

    try:
        if include_mm_ranking:
            mode_iter = tqdm(MODES, desc="MM crawl", unit="mode") if HAS_TQDM else MODES
            for mode in mode_iter:
                if is_stop_requested():
                    logger.info("Stop requested, interrupt MM ranking crawl.")
                    break
                if consume_skip_request():
                    logger.warning("Skip requested, skip MM ranking mode %d.", mode)
                    continue
                task = f"mm_global_mode_{mode}"
                logger.info("MM ranking mode %d start...", mode)
                try:
                    rows = fetch_mm_global_mode(session, mode, mm_limit)
                    for item in rows:
                        uid = str(_safe_int(item.get("uid"), 0))
                        if uid != "0":
                            mm_uids.add(uid)
                    row_stats = save_mm_ranking_rows(
                        mode=mode,
                        rows=rows,
                        crawl_time=crawl_time,
                        source="mm_global",
                    )
                    stats["mm_rows"] += len(rows)
                    stats["mm_modes_ok"] += 1
                    update_mm_crawl_status(task, True, state=row_stats)
                    logger.info(
                        "MM ranking mode %d done: fetched=%d saved=%d",
                        mode,
                        row_stats.get("rows", 0),
                        row_stats.get("saved_rows", 0),
                    )
                    if HAS_TQDM:
                        mode_iter.set_postfix_str(
                            f"rows={stats['mm_rows']} ok={stats['mm_modes_ok']} fail={stats['mm_modes_fail']}"
                        )
                        tqdm.write(
                            colorize(
                                f"[MM][mode {mode}] fetched={row_stats.get('rows', 0)} saved={row_stats.get('saved_rows', 0)}",
                                "96",
                            )
                        )
                    else:
                        print(
                            colorize(
                                f"[MM][mode {mode}] fetched={row_stats.get('rows', 0)} saved={row_stats.get('saved_rows', 0)}",
                                "96",
                            )
                        )
                    time.sleep(0.2)
                except SkipCurrentTask as e:
                    update_mm_crawl_status(task, False, error="skipped_by_user", state={"mode": mode})
                    logger.warning("Skip requested, %s skipped.", e)
                    continue
                except Exception as e:
                    stats["mm_modes_fail"] += 1
                    update_mm_crawl_status(task, False, error=str(e), state={"mode": mode})
                    logger.warning("MM global crawl failed for mode %d: %s", mode, e)
                    if HAS_TQDM:
                        mode_iter.set_postfix_str(
                            f"rows={stats['mm_rows']} ok={stats['mm_modes_ok']} fail={stats['mm_modes_fail']}"
                        )
                        tqdm.write(colorize(f"[MM][mode {mode}] failed: {e}", "91"))
                    else:
                        print(colorize(f"[MM][mode {mode}] failed: {e}", "91"))
        else:
            mm_top_uids = get_recent_mm_tracked_uids(limit_per_mode=mm_limit)
            mm_uids.update(mm_top_uids)
            stats["mmr_uid_from_mm_top"] = len(mm_top_uids)

        if include_mm_ranking:
            stats["mmr_uid_from_mm_top"] = len(mm_uids)

        if include_exp_ranking_uids:
            exp_top_uids = get_recent_exp_tracked_uids(limit_per_mode=max(1, int(exp_limit or mm_limit)))
            before = len(mm_uids)
            mm_uids.update(exp_top_uids)
            stats["mmr_uid_from_exp_top"] = len(exp_top_uids)
            logger.info(
                "MMR uid pool add from EXP top: raw=%d merged=%d(+%d)",
                len(exp_top_uids),
                len(mm_uids),
                len(mm_uids) - before,
            )

        config_players = load_player_config()
        manual_uids = {str(uid).strip() for uid in config_players if str(uid).strip().isdigit()}
        before_manual = len(mm_uids)
        mm_uids.update(manual_uids)
        stats["mmr_uid_from_manual"] = len(manual_uids)
        logger.info(
            "MMR uid pool add from manual config: raw=%d merged=%d(+%d)",
            len(manual_uids),
            len(mm_uids),
            len(mm_uids) - before_manual,
        )

        if seed_uids:
            valid_seed = {str(uid).strip() for uid in seed_uids if str(uid).strip().isdigit()}
            before_seed = len(mm_uids)
            mm_uids.update(valid_seed)
            stats["mmr_uid_from_seed"] = len(valid_seed)
            logger.info(
                "MMR uid pool add from seed: raw=%d merged=%d(+%d)",
                len(valid_seed),
                len(mm_uids),
                len(mm_uids) - before_seed,
            )

        stats["mmr_uid_pool"] = len(mm_uids)
        logger.info(
            "MMR uid pool summary: total=%d mm_top=%d exp_top=%d manual=%d seed=%d",
            stats["mmr_uid_pool"],
            stats["mmr_uid_from_mm_top"],
            stats["mmr_uid_from_exp_top"],
            stats["mmr_uid_from_manual"],
            stats["mmr_uid_from_seed"],
        )

        if include_mmr:
            uid_list = sorted(mm_uids)
            if mmr_min_interval_minutes and mmr_min_interval_minutes > 0 and uid_list:
                db_manager = DatabaseManager()
                cursor = db_manager.get_connection().cursor()
                cutoff = datetime.now() - timedelta(minutes=int(mmr_min_interval_minutes))
                cursor.execute(
                    """
                    SELECT DISTINCT uid
                    FROM player_mmr_samples
                    WHERE crawl_time >= ?
                      AND uid IS NOT NULL
                      AND uid != ''
                    """,
                    (cutoff,),
                )
                recent_uids = {str(row[0]) for row in cursor.fetchall() if row and row[0]}
                filtered = [uid for uid in uid_list if uid not in recent_uids]
                stats["mmr_users_skipped_recent"] = len(uid_list) - len(filtered)
                if stats["mmr_users_skipped_recent"] > 0:
                    logger.info(
                        "MMR request de-dup: skip %d users sampled within %d minutes.",
                        stats["mmr_users_skipped_recent"],
                        mmr_min_interval_minutes,
                    )
                uid_list = filtered
            total_users = len(uid_list)
            logger.info("MMR sync start: users=%d", total_users)
            mmr_iter = tqdm(uid_list, desc="MMR抓取", unit="玩家") if HAS_TQDM else uid_list
            for idx, uid in enumerate(mmr_iter, start=1):
                if is_stop_requested():
                    logger.info("Stop requested, interrupt MMR crawl.")
                    break
                if consume_skip_request():
                    logger.warning("Skip requested, skip current MMR uid %s.", uid)
                    continue
                try:
                    player_rows = fetch_player_mmr_from_api(session, uid)
                    source = "ranking_player_all"
                    if not player_rows:
                        player_rows = fetch_player_mmr_from_page(session, uid)
                        source = "page_profile"
                    if not player_rows:
                        stats["mmr_users_fail"] += 1
                        continue
                    for item in player_rows:
                        if is_stop_requested():
                            break
                        if consume_skip_request():
                            raise SkipCurrentTask(f"mmr uid {uid}")
                        mode = _safe_int(item.get("mode"), -1)
                        mmr = _safe_int(item.get("mmr"), -1)
                        if mode < 0 or mmr < 0:
                            continue
                        name = item.get("name")
                        if name:
                            player_id = resolve_player_identity(name, crawl_time, uid=uid)
                        else:
                            player_id = get_player_id_by_uid(uid)
                        sample_op = save_player_mmr_sample(
                            player_id=player_id,
                            uid=uid,
                            mode=mode,
                            mmr=mmr,
                            mm_rank=item.get("mm_rank"),
                            name=name,
                            crawl_time=crawl_time,
                            source=source,
                        )
                        if sample_op in ("new", "diff_insert", "same_insert"):
                            stats["mmr_samples"] += 1
                        if sample_op == "new":
                            stats["mmr_sample_new"] += 1
                        elif sample_op == "diff_insert":
                            stats["mmr_sample_diff_insert"] += 1
                        elif sample_op == "same_insert":
                            stats["mmr_sample_same_insert"] += 1
                        elif sample_op == "update":
                            stats["mmr_sample_update"] += 1
                    stats["mmr_users_ok"] += 1
                    time.sleep(0.1)
                except SkipCurrentTask as e:
                    logger.warning("Skip requested, %s skipped.", e)
                    continue
                except Exception as e:
                    stats["mmr_users_fail"] += 1
                    logger.warning("MMR crawl failed for uid=%s: %s", uid, e)
                    if HAS_TQDM:
                        tqdm.write(colorize(f"[MMR][{idx}/{total_users}] uid={uid} failed: {e}", "91"))
                    else:
                        print(colorize(f"[MMR][{idx}/{total_users}] uid={uid} failed: {e}", "91"))
                if HAS_TQDM:
                    mmr_iter.set_postfix_str(
                        f"ok={stats['mmr_users_ok']} fail={stats['mmr_users_fail']} samples={stats['mmr_samples']}"
                    )
                if idx == 1 or idx % 20 == 0 or idx == total_users:
                    line = (
                        f"[MMR][{idx}/{total_users}] "
                        f"ok={stats['mmr_users_ok']} fail={stats['mmr_users_fail']} samples={stats['mmr_samples']}"
                    )
                    if HAS_TQDM:
                        tqdm.write(colorize(line, "93"))
                    else:
                        print(colorize(line, "93"))
                    logger.info(
                        "MMR progress: %d/%d users, ok=%d fail=%d samples=%d",
                        idx,
                        total_users,
                        stats["mmr_users_ok"],
                        stats["mmr_users_fail"],
                        stats["mmr_samples"],
                    )

            update_mm_crawl_status(
                "mmr_batch",
                success=(stats["mmr_users_fail"] == 0),
                error=None if stats["mmr_users_fail"] == 0 else f"failed={stats['mmr_users_fail']}",
                state=stats,
            )

        # Notify if all MM modes failed
        if include_mm_ranking and stats.get("mm_modes_fail", 0) > 0 and stats.get("mm_modes_ok", 0) == 0:
            notify_user("MM排行爬取失败", "所有模式的MM排行均爬取失败，可能需要更新Cookie/Token", "error")

        logger.info("MM sync done: %s", stats)
        return stats
    finally:
        _release_mm_run_lock()


def save_to_database(mode, df, crawl_time):
    """Smart-save leaderboard rows from DataFrame and print operation stats."""
    if df.empty:
        return

    stats = {'new': 0, 'diff_insert': 0, 'same_insert': 0, 'update': 0, 'update_fill': 0}

    for _, row in df.iterrows():
        player_id = resolve_player_identity(row['name'], crawl_time, row.get('player_id'))
        if player_id is not None:
            op = save_player_ranking_record(
                player_id=player_id,
                uid=row.get('player_id'),
                mode=mode,
                rank=row['rank'],
                name=row['name'],
                lv=row['lv'],
                exp=row['exp'],
                acc=row['acc'],
                combo=row['combo'],
                pc=row['pc'],
                crawl_time=crawl_time,
                source='leaderboard'  # 排行榜数据
            )
            stats[op] += 1

    total = sum(stats.values())
    msg = (f"模式 {mode} 数据处理完成: 总计 {total} 条 | "
           f"{colorize('首次', '92')}: {stats['new']} | "
           f"{colorize('变化插入', '93')}: {stats['diff_insert']} | "
           f"{colorize('相同插入', '94')}: {stats['same_insert']} | "
           f"{colorize('时间更新', '95')}: {stats['update']} | "
           f"{colorize('LV填充', '96')}: {stats['update_fill']}")
    print(msg)
    logger.info("模式 %d 数据统计: new=%d, diff=%d, same=%d, update=%d, fill=%d",
                mode, stats['new'], stats['diff_insert'], stats['same_insert'], stats['update'], stats['update_fill'])


def save_player_profile_to_database(player_data, crawl_time, player_identifier):
    """Smart-save player profile ranking rows and update crawl status."""
    if not player_data:
        return False

    db_manager = DatabaseManager()
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    success = False
    stats = {'new': 0, 'diff_insert': 0, 'same_insert': 0, 'update': 0, 'update_fill': 0}

    try:
        for data in player_data:
            player_id = resolve_player_identity(data['name'], crawl_time, player_identifier)
            if player_id is not None:
                op = save_player_ranking_record(
                    player_id=player_id,
                    uid=player_identifier,
                    mode=data['mode'],
                    rank=data['rank'],
                    name=data['name'],
                    lv=data['lv'],
                    exp=data['exp'],
                    acc=data['acc'],
                    combo=data['combo'],
                    pc=data['pc'],
                    crawl_time=crawl_time,
                    source='profile'  # 个人页数据
                )
                stats[op] += 1
                success = True

        # Update crawl status after processing this player.
        cursor.execute('''
            INSERT OR REPLACE INTO player_crawl_status
            (player_identifier, last_crawled, crawl_count, success_count, last_error)
            VALUES (?, ?,
                COALESCE((SELECT crawl_count FROM player_crawl_status WHERE player_identifier = ?), 0) + 1,
                COALESCE((SELECT success_count FROM player_crawl_status WHERE player_identifier = ?), 0) + ?,
                NULL)
        ''', (player_identifier, crawl_time, player_identifier, player_identifier, 1 if success else 0))
        conn.commit()

        total = sum(stats.values())
        msg = (f"玩家 {player_identifier} 数据处理完成: 总计 {total} 条 | "
               f"{colorize('首次', '92')}: {stats['new']} | "
               f"{colorize('变化插入', '93')}: {stats['diff_insert']} | "
               f"{colorize('相同插入', '94')}: {stats['same_insert']} | "
               f"{colorize('时间更新', '95')}: {stats['update']} | "
               f"{colorize('LV填充', '96')}: {stats['update_fill']}")
        print(msg)
        logger.info("玩家 %s 数据统计: new=%d, diff=%d, same=%d, update=%d, fill=%d",
                    player_identifier, stats['new'], stats['diff_insert'], stats['same_insert'], stats['update'], stats['update_fill'])
        return success

    except Exception as e:
        logger.error("保存玩家 %s 数据到数据库失败: %s", player_identifier, e)
        # Persist failed crawl status for this player.
        cursor.execute('''
            INSERT OR REPLACE INTO player_crawl_status
            (player_identifier, last_crawled, crawl_count, success_count, last_error)
            VALUES (?, ?,
                COALESCE((SELECT crawl_count FROM player_crawl_status WHERE player_identifier = ?), 0) + 1,
                COALESCE((SELECT success_count FROM player_crawl_status WHERE player_identifier = ?), 0),
                ?)
        ''', (player_identifier, crawl_time, player_identifier, player_identifier, str(e)))
        conn.commit()
        return False


def check_excel_file_integrity(filename):
    """Check whether an Excel file is structurally valid and readable."""
    try:
        if not os.path.exists(filename):
            return False
        with open(filename, 'rb') as f:
            header = f.read(8)
            if header[:4] != b'PK\x03\x04':
                return False
        wb = load_workbook(filename, read_only=True)
        sheetnames = wb.sheetnames
        wb.close()
        return bool(sheetnames)
    except Exception as e:
        logger.warning("Excel文件完整性检查失败: %s - %s", filename, e)
        return False


def repair_excel_file(filename):
    """Attempt to repair a corrupted Excel file."""
    try:
        backup_name = f"{filename}.backup.{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        import shutil
        shutil.copy2(filename, backup_name)
        logger.info("已创建备份文件: %s", backup_name)

        try:
            xl = pd.ExcelFile(filename)
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                for sheet_name in xl.sheet_names:
                    df = pd.read_excel(filename, sheet_name=sheet_name)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            logger.info("成功修复Excel文件: %s", filename)
            return True
        except Exception as e:
            logger.error("使用pandas修复失败: %s", e)
            try:
                wb = load_workbook(filename)
                wb.save(filename)
                logger.info("使用openpyxl修复成功: %s", filename)
                return True
            except Exception as e2:
                logger.error("使用openpyxl修复也失败: %s", e2)
                return False
    except Exception as e:
        logger.error("修复Excel文件过程中发生错误: %s", e)
        return False


def import_mode_data(mode):
    """Import one mode historical data from Excel into database."""
    global stop_requested

    thread_id = threading.get_ident()
    db_manager = DatabaseManager()
    conn = db_manager.get_connection(thread_id)

    conn.execute("PRAGMA synchronous = OFF")
    conn.execute("PRAGMA journal_mode = MEMORY")
    conn.execute("PRAGMA cache_size = 10000")
    conn.execute("PRAGMA temp_store = MEMORY")

    player_alias_cache = {}

    filename = get_excel_filename(mode)
    if not os.path.exists(filename):
        logger.warning("模式 %d 的Excel文件不存在: %s", mode, filename)
        return 0

    if not check_excel_file_integrity(filename):
        logger.warning("模式 %d 的Excel文件可能已损坏: %s", mode, filename)
        if repair_excel_file(filename):
            logger.info("文件修复成功，继续导入")
        else:
            logger.error("文件修复失败，跳过模式 %d", mode)
            return 0

    cursor = conn.cursor()
    cursor.execute(
        "SELECT last_import_time FROM import_metadata WHERE mode = ?",
        (mode,)
    )
    result = cursor.fetchone()
    last_import_time = result[0] if result else None

    try:
        xl = pd.ExcelFile(filename)
        sheet_names = xl.sheet_names
    except Exception as e:
        logger.error("打开Excel文件失败: %s", e)
        try:
            wb = load_workbook(filename)
            sheet_names = wb.sheetnames
        except Exception as e2:
            logger.error("两种方式都无法打开Excel文件: %s", e2)
            return 0

    sheet_names = [s for s in sheet_names if s.startswith(f"mode_{mode}_")]

    sheet_times = []
    for sheet_name in sheet_names:
        try:
            time_str = sheet_name.replace(f"mode_{mode}_", "")
            sheet_time = datetime.strptime(time_str, "%Y-%m-%d_%H-%M")
            sheet_times.append((sheet_name, sheet_time))
        except:
            continue

    sheet_times.sort(key=lambda x: x[1])

    if HAS_TQDM:
        mode_pbar = tqdm(
            sheet_times,
            desc=f"模式 {mode}",
            position=mode + 1,
            leave=False,
            unit="row"
        )
    else:
        mode_pbar = sheet_times
        print(f"开始处理模式 {mode}，共 {len(sheet_times)} 个表...")

    imported_count = 0
    batch_size = 50
    batch_data = []

    for i, (sheet_name, sheet_time) in enumerate(mode_pbar):
        with stop_lock:
            if stop_requested:
                logger.info("Mode %d import interrupted, imported %d rows.", mode, imported_count)
                break

        if last_import_time and sheet_time <= datetime.strptime(last_import_time, "%Y-%m-%d %H:%M:%S"):
            if HAS_TQDM:
                mode_pbar.update(1)
            continue

        try:
            df = pd.read_excel(filename, sheet_name=sheet_name)
            if df.empty:
                if HAS_TQDM:
                    mode_pbar.update(1)
                continue

            for _, row in df.iterrows():
                name = row['name']
                if name in player_alias_cache:
                    player_id = player_alias_cache[name]
                else:
                    player_id = resolve_player_identity(name, sheet_time)
                    player_alias_cache[name] = player_id

                if player_id is not None:
                    batch_data.append((
                        player_id, None, mode, row['rank'], name, row['lv'], row['exp'],
                        row['acc'], row['combo'], row['pc'], sheet_time
                    ))

            if batch_data and (len(batch_data) >= 1000 or i == len(sheet_times) - 1):
                cursor.executemany('''
                INSERT OR IGNORE INTO player_rankings
                (player_id, uid, mode, rank, name, lv, exp, acc, combo, pc, crawl_time)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', batch_data)
                imported_count += len(batch_data)
                batch_data = []

            if (i + 1) % batch_size == 0 or i == len(sheet_times) - 1:
                cursor.execute(
                    "UPDATE import_metadata SET last_import_time = ? WHERE mode = ?",
                    (sheet_time.strftime("%Y-%m-%d %H:%M:%S"), mode)
                )
                conn.commit()

            if HAS_TQDM:
                mode_pbar.set_postfix_str(f"已导入 {imported_count}")

        except Exception as e:
            logger.error("导入模式 %d 表 %s 时出错: %s", mode, sheet_name, e)
            conn.rollback()
            continue

    if batch_data:
        cursor.executemany('''
        INSERT OR IGNORE INTO player_rankings
        (player_id, uid, mode, rank, name, lv, exp, acc, combo, pc, crawl_time)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', batch_data)
        imported_count += len(batch_data)
        conn.commit()

    conn.execute("PRAGMA synchronous = NORMAL")
    conn.execute("PRAGMA journal_mode = WAL")
    conn.commit()

    db_manager.close_connection(thread_id)

    if HAS_TQDM:
        mode_pbar.close()

    return imported_count


def import_historical_data():
    """Import historical data from all mode Excel files into database."""
    if HAS_TQDM:
        main_pbar = tqdm(total=len(MODES), desc="总体进度", position=0)
    else:
        print("开始导入历史数据...")

    for mode in MODES:
        try:
            result = import_mode_data(mode)
            if HAS_TQDM:
                main_pbar.update(1)
                main_pbar.set_postfix_str(f"模式 {mode} 完成: {result} 条记录")
            else:
                print(f"模式 {mode} 完成: {result} 条记录")
        except Exception as e:
            logger.error("模式 %d 导入失败: %s", mode, e)
            if HAS_TQDM:
                main_pbar.update(1)
                main_pbar.set_postfix_str(f"模式 {mode} 失败: {e}")
            else:
                print(f"模式 {mode} 失败: {e}")

    if HAS_TQDM:
        main_pbar.close()

    if not HAS_TQDM:
        print("历史数据导入完成")


def load_player_config():
    """Load players.txt and return configured player IDs."""
    players = []
    if not os.path.exists(PLAYER_CONFIG_FILE):
        logger.info("玩家配置文件不存在，创建空文件: %s", PLAYER_CONFIG_FILE)
        with open(PLAYER_CONFIG_FILE, 'w', encoding='utf-8') as f:
            f.write("# 每行一个玩家ID（必须是数字）\n")
            f.write("# 例如:\n")
            f.write("# 923177\n")
            f.write("# 123456\n")
        return players
    try:
        with open(PLAYER_CONFIG_FILE, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    players.append(line)
        logger.info("Loaded %d players from config file.", len(players))
        return players
    except Exception as e:
        logger.error("加载玩家配置文件失败: %s", e)
        return []


def add_players_to_queue(players):
    """Add players into crawl queue with deduplication."""
    added = 0
    with _player_set_lock:
        for player in players:
            if player not in _player_set:
                _player_set.add(player)
                player_queue.put(player)
                added += 1
    if added > 0:
        logger.info("Added %d new players into crawl queue.", added)


def get_players_from_leaderboard(df_list):
    """Extract all unique player IDs from leaderboard DataFrames."""
    players = set()
    for df in df_list:
        if not df.empty and 'player_id' in df.columns:
            for player_id in df['player_id']:
                if player_id:
                    players.add(player_id)
    return list(players)


def run_player_crawler():
    """Run player profile crawler synchronously."""
    global player_crawl_in_progress, last_player_crawl_time

    with player_crawl_lock:
        if player_crawl_in_progress:
            logger.info("玩家爬取器已在运行，跳过")
            return
        player_crawl_in_progress = True

    try:
        logger.info("开始玩家个人主页爬取周期")
        last_player_crawl_time = datetime.now()

        session = requests.Session()
        session.cookies.update(COOKIES)
        session.headers.update(HEADERS)
        session.mount('https://', requests.adapters.HTTPAdapter(max_retries=3))

        with _player_set_lock:
            total_players = len(_player_set)
        if total_players == 0:
            logger.info("爬取队列为空，跳过")
            return

        logger.info("开始爬取 %d 个玩家的个人主页数据", total_players)

        successful_crawls = 0
        failed_crawls = 0

        if HAS_TQDM:
            pbar = tqdm(total=total_players, desc="玩家主页爬取", unit="玩家")
        else:
            print(f"开始爬取 {total_players} 个玩家的个人主页数据...")

        while True:
            if is_stop_requested():
                logger.info("Player crawler interrupted by stop request.")
                break

            try:
                player_identifier = player_queue.get_nowait()
            except queue.Empty:
                break

            # Remove from dedupe set once popped from queue.
            with _player_set_lock:
                _player_set.discard(player_identifier)

            try:
                player_data = crawl_player_profile(session, player_identifier)
                crawl_time = datetime.now()

                if consume_skip_request():
                    logger.warning("Skip requested, skip player %s.", player_identifier)
                    continue

                if player_data and save_player_profile_to_database(player_data, crawl_time, player_identifier):
                    successful_crawls += 1
                else:
                    failed_crawls += 1

                if HAS_TQDM:
                    pbar.set_postfix_str(f"成功: {successful_crawls}, 失败: {failed_crawls}")

                time.sleep(3)

            except Exception as e:
                logger.error("Error while processing player %s: %s", player_identifier, e)
                failed_crawls += 1
                if HAS_TQDM:
                    pbar.set_postfix_str(f"成功: {successful_crawls}, 失败: {failed_crawls}")
            finally:
                if HAS_TQDM:
                    _ = pbar.update(1)
                player_queue.task_done()

        if HAS_TQDM:
            pbar.close()

        logger.info("玩家个人主页爬取完成: 成功 %d, 失败 %d", successful_crawls, failed_crawls)

    except Exception as e:
        logger.error("玩家爬取周期发生错误: %s", e)
    finally:
        with player_crawl_lock:
            player_crawl_in_progress = False


def git_check_updates():
    """Check remote Git repo updates for .db/.xlsx data files only."""
    try:
        if not os.path.exists(os.path.join(GIT_REPO_PATH, '.git')):
            logger.info("当前目录不是Git仓库，跳过Git更新检查")
            return False

        original_cwd = os.getcwd()
        os.chdir(GIT_REPO_PATH)

        result = subprocess.run(["git", "remote", "-v"], capture_output=True, text=True)
        if not result.stdout.strip():
            logger.info("未配置Git远程仓库，跳过更新检查")
            os.chdir(original_cwd)
            return False

        result = subprocess.run(["git", "fetch", "origin"], capture_output=True, text=True)
        if result.returncode != 0:
            logger.warning("Git fetch失败: %s", result.stderr)
            os.chdir(original_cwd)
            return False

        result = subprocess.run(
            ["git", "diff", "--name-only", "HEAD", "origin/main", "--", "*.db", "*.xlsx"],
            capture_output=True, text=True
        )

        os.chdir(original_cwd)

        if result.stdout.strip():
            updated_files = result.stdout.strip().split('\n')
            logger.info("发现远程更新文件: %s", updated_files)
            return True
        else:
            logger.info("远程仓库没有 .db 或 xlsx 文件的更新")
            return False

    except subprocess.CalledProcessError as e:
        logger.warning("Git检查更新失败: %s", e)
        return False
    except Exception as e:
        logger.warning("Git检查更新发生意外错误: %s", e)
        return False


def git_pull_data_files():
    """Pull data files from remote Git repo with interactive confirmation."""
    try:
        if not os.path.exists(os.path.join(GIT_REPO_PATH, '.git')):
            logger.info("当前目录不是Git仓库，跳过Git拉取")
            return False

        print("\n" + "="*60)
        print("检测到远程仓库有更新！")
        print("更新文件包括: catch.xlsx, key.xlsx, malody_rankings.db 等数据文件")
        print("这些更新将覆盖您本地的数据文件。")
        print("="*60)
        print("您有10秒时间决定是否拉取更新：")
        print("  - 输入 'y' 或 'yes' 确认拉取")
        print("  - 输入 'n' 或 'no' 跳过拉取")
        print("  - 10秒内无响应将自动跳过")
        print("="*60)

        try:
            import select
            import sys
            print("请确认是否拉取更新 (y/n, 10秒超时): ", end='', flush=True)
            start_time = time.time()
            while time.time() - start_time < 10:
                if select.select([sys.stdin], [], [], 0.1)[0]:
                    user_input = sys.stdin.readline().strip().lower()
                    if user_input in ['y', 'yes']:
                        print("确认拉取更新...")
                        break
                    elif user_input in ['n', 'no']:
                        print("跳过拉取更新。")
                        return False
                    else:
                        print("无效输入，请输入 y 或 n: ", end='', flush=True)
                time.sleep(0.1)
            else:
                print("\n10秒超时，自动跳过拉取更新。")
                return False
        except ImportError:
            print("请确认是否拉取更新 (y/n, 10秒超时): ", end='', flush=True)
            user_input = input()
            if user_input.lower() in ['y', 'yes']:
                print("确认拉取更新...")
            else:
                print("跳过拉取更新。")
                return False

        original_cwd = os.getcwd()
        os.chdir(GIT_REPO_PATH)

        DatabaseManager().close_connection()

        excel_files = [get_excel_filename(mode) for mode in MODES]
        files_to_pull = [f for f in excel_files if os.path.exists(f)]
        files_to_pull.append(DB_FILE)

        # Pull data files one-by-one from origin/main so code files remain untouched.
        success = True
        for file in files_to_pull:
            result = subprocess.run(
                ["git", "checkout", "origin/main", "--", file],
                capture_output=True, text=True
            )
            if result.returncode == 0:
                logger.info("已拉取文件: %s", file)
            else:
                logger.warning("拉取文件 %s 失败: %s", file, result.stderr)
                success = False

        os.chdir(original_cwd)
        return success
    except subprocess.CalledProcessError as e:
        logger.warning("Git拉取文件失败: %s", e)
        return False
    except Exception as e:
        logger.warning("Git拉取文件发生意外错误: %s", e)
        return False


def git_add_commit_push(has_changes=True):
    """Auto add/commit/push Git changes."""
    if not has_changes:
        logger.info("所有模式均无数据变化，跳过Git推送")
        return True

    try:
        if not os.path.exists(os.path.join(GIT_REPO_PATH, '.git')):
            logger.info("当前目录不是Git仓库，跳过Git推送")
            return True

        original_cwd = os.getcwd()
        os.chdir(GIT_REPO_PATH)

        result = subprocess.run(["git", "remote", "-v"], capture_output=True, text=True)
        if not result.stdout.strip():
            logger.info("未配置Git远程仓库，跳过推送")
            os.chdir(original_cwd)
            return True

        excel_files = [get_excel_filename(mode) for mode in MODES]
        files_to_add = [f for f in excel_files if os.path.exists(f)]
        files_to_add.append(DB_FILE)

        for file in files_to_add:
            result = subprocess.run(["git", "add", file], capture_output=True, text=True)
            if result.returncode != 0:
                logger.warning("添加文件 %s 失败: %s", file, result.stderr)

        result = subprocess.run(["git", "status", "--porcelain"], capture_output=True, text=True)
        if not result.stdout.strip():
            logger.info("没有文件更改，跳过Git提交")
            os.chdir(original_cwd)
            return True

        result = subprocess.run(["git", "commit", "-m", GIT_COMMIT_MESSAGE], capture_output=True, text=True)
        if result.returncode == 0:
            logger.info("Git提交成功: %s", GIT_COMMIT_MESSAGE)
        else:
            logger.warning("Git提交失败: %s", result.stderr)
            os.chdir(original_cwd)
            return False

        result = subprocess.run(["git", "push"], capture_output=True, text=True)
        if result.returncode == 0:
            logger.info("Git推送成功")
            success = True
        else:
            logger.warning("Git推送失败: %s", result.stderr)
            success = False

        os.chdir(original_cwd)
        return success
    except subprocess.CalledProcessError as e:
        logger.warning("Git操作失败: %s", e)
        return False
    except Exception as e:
        logger.warning("Git操作发生意外错误: %s", e)
        return False


def check_data_changed(mode, df):
    """Check whether current data equals latest Excel snapshot."""
    filename = get_excel_filename(mode)
    sheet_name = f"mode_{mode}"

    if not os.path.exists(filename):
        return True

    try:
        wb = load_workbook(filename)
        if sheet_name not in wb.sheetnames:
            return True

        sub_sheets = [s for s in wb.sheetnames if s.startswith(f"{sheet_name}_")]
        if not sub_sheets:
            return True

        latest_sheet = None
        latest_time = None
        for s in sub_sheets:
            try:
                dt_str = s.replace(f"{sheet_name}_", "")
                dt = datetime.strptime(dt_str, "%Y-%m-%d_%H-%M")
                if latest_time is None or dt > latest_time:
                    latest_time = dt
                    latest_sheet = s
            except:
                continue

        if latest_sheet:
            df_prev = pd.read_excel(filename, sheet_name=latest_sheet)
            if not df_prev.empty and df_prev.equals(df):
                return False
    except Exception as e:
        logger.error("检查数据变化时出错: %s", e)
        return True

    return True


def run_crawler_cycle(
    crawl_players=False,
    crawl_leaderboard_players=False,
    save_excel=False,
    push_to_git=False,
    mm_sync=False,
    mm_limit=DEFAULT_MM_LIMIT,
    skip_mm_ranking=False,
    skip_mmr=False,
    ranking_source="page",
    ranking_limit=DEFAULT_MM_LIMIT,
):
    """
    运行一个爬取周期。

    Args:
        crawl_players: 是否启动玩家爬虫（消费队列中的玩家ID，包含配置文件中的玩家）
        crawl_leaderboard_players: 是否将当前排行榜中的玩家ID加入队列
        save_excel: 是否保存数据到 Excel 文件
        push_to_git: 是否推送数据到 Git 仓库
        mm_sync: 是否在本轮执行 MM/MMR 同步
        mm_limit: 每个模式抓取的 MM 榜单上限
    """
    # Reload configured players at the beginning of each cycle.
    if crawl_players:
        config_players = load_player_config()
        if config_players:
            add_players_to_queue(config_players)
            logger.info("周期开始：已重新加载 %d 个配置文件玩家到队列", len(config_players))

    try:
        if push_to_git and git_check_updates():
            logger.info("检测到远程仓库有更新，正在拉取数据文件...")
            if git_pull_data_files():
                logger.info("数据文件拉取完成，重新初始化数据库...")
                DatabaseManager().close_connection()
                init_database()
            else:
                logger.warning("数据文件拉取失败，继续使用本地数据")
        else:
            logger.info("未检测到远程更新或Git不可用，继续使用本地数据")
    except Exception as e:
        logger.warning("Git更新检查失败，继续使用本地数据: %s", e)

    session = requests.Session()
    session.cookies.update(COOKIES)
    session.headers.update(HEADERS)
    session.mount('https://', requests.adapters.HTTPAdapter(max_retries=3))

    start_time = datetime.now()
    logger.info("=" * 50)
    logger.info("开始爬取周期: %s", start_time)

    has_changes = False
    all_dfs = []

    for mode in MODES:
        if is_stop_requested():
            logger.info("Stop requested, interrupt ranking crawl cycle.")
            break
        if consume_skip_request():
            logger.warning("Skip requested, skip current ranking mode %d.", mode)
            continue

        try:
            logger.info("Processing ranking mode: %d", mode)
            mm_rows: List[Dict[str, Any]] = []
            if ranking_source == "newapi":
                df, mm_rows = crawl_mode_player_newapi(session, mode, ranking_limit)
            else:
                df = crawl_mode_player(session, mode)
            all_dfs.append(df)

            if consume_skip_request():
                raise SkipCurrentTask(f"ranking mode {mode}")

            if df.empty:
                logger.warning("Mode %d returned empty ranking data, skip.", mode)
                notify_user("排行榜数据为空", f"模式 {mode} 返回空排行榜数据，可能需要更新Cookie/Token", "warning")
                continue

            if save_excel and not check_data_changed(mode, df):
                logger.info("Mode %d data unchanged, skip Excel save.", mode)

            crawl_time = datetime.now()

            if save_excel and check_data_changed(mode, df):
                save_data_to_excel(mode, df, crawl_time)

            if consume_skip_request():
                raise SkipCurrentTask(f"ranking mode {mode}")

            save_to_database(mode, df, crawl_time)
            if ranking_source == "newapi" and mm_rows:
                mm_row_stats = save_mm_ranking_rows(
                    mode=mode,
                    rows=mm_rows,
                    crawl_time=crawl_time,
                    source="newapi_ranking_global",
                )
                update_mm_crawl_status(
                    task=f"mm_global_mode_{mode}",
                    success=True,
                    state={"source": "newapi_ranking_global", **mm_row_stats},
                )
            has_changes = True

            time.sleep(3)
        except SkipCurrentTask as e:
            logger.warning("Skip requested, %s skipped.", e)
            continue
        except Exception as e:
            logger.exception("Unexpected error while processing mode %d", mode)

    # Seed MM/MMR sync with players seen in this cycle's leaderboard snapshots.
    mm_seed_uids: Set[str] = set(get_players_from_leaderboard(all_dfs)) if all_dfs else set()

    if mm_sync and not is_stop_requested():
        logger.info("Start MM/MMR sync...")
        if consume_skip_request():
            logger.warning("Skip requested, skip MM/MMR sync in this cycle.")
        else:
            try:
                include_mm_ranking = not skip_mm_ranking
                if ranking_source == "newapi" and include_mm_ranking:
                    # newapi ranking already includes MM rows; skip duplicate fetch in mm_sync.
                    logger.info("Ranking source is newapi, skip duplicate MM ranking fetch in mm_sync.")
                    include_mm_ranking = False
                run_mm_sync_cycle(
                    mm_limit=mm_limit,
                    include_mm_ranking=include_mm_ranking,
                    include_mmr=not skip_mmr,
                    include_exp_ranking_uids=True,
                    exp_limit=ranking_limit if ranking_source == "newapi" else mm_limit,
                    seed_uids=mm_seed_uids,
                )
            except Exception as e:
                logger.warning("MM/MMR sync failed, continue with existing flow: %s", e)

    if crawl_players and crawl_leaderboard_players and all_dfs:
        leaderboard_players = get_players_from_leaderboard(all_dfs)
        if leaderboard_players:
            add_players_to_queue(leaderboard_players)
            logger.info("从排行榜添加了 %d 个玩家到爬取队列", len(leaderboard_players))

    # Run player-profile crawl synchronously in current cycle.
    if crawl_players:
        logger.info("同步执行玩家主页爬取...")
        run_player_crawler()
    else:
        logger.info("玩家爬取已禁用，跳过")

    # Push to Git only when explicitly enabled by CLI option.
    if push_to_git:
        try:
            _ = git_add_commit_push(has_changes)
        except Exception as e:
            logger.warning("Git推送失败，但数据已保存在本地: %s", e)
    else:
        logger.info("Git推送已禁用，数据仅保存在本地")

    # Summary notification
    total_modes = len(all_dfs)
    empty_modes = sum(1 for df in all_dfs if df.empty)
    successful_modes = total_modes - empty_modes
    if empty_modes > 0:
        if successful_modes == 0:
            notify_user(
                "排行榜爬取失败",
                f"所有 {total_modes} 个模式均返回空数据，可能需要更新Cookie/Token",
                "error"
            )
        else:
            notify_user(
                "排行榜爬取部分失败",
                f"{empty_modes}/{total_modes} 个模式返回空数据",
                "warning"
            )
    else:
        notify_user("排行榜爬取完成", f"成功爬取 {total_modes} 个模式的数据", "info")

    end_time = datetime.now()
    duration = (end_time - start_time).total_seconds()
    logger.info("Crawl cycle completed, duration: %.2fs", duration)
    logger.info("=" * 50)


def optimize_database():
    """
    优化数据库：按照新规则清理旧数据，只保留必要记录。
    对于每个玩家每个模式，按时间排序：
      - 以保存逻辑一致的核心字段（rank/name/exp/acc/combo/pc）划分稳定段。
      - 对稳定段仅保留边界两条记录（start/end），删除中间重复记录。
    最终结果：每个稳定段最多保留两条边界记录（与区间模型保存逻辑一致）。
    打印优化前后的记录数和数据库文件大小，并计算优化率。
    """
    logger.info("开始数据库优化...")
    db_manager = DatabaseManager()
    conn = db_manager.get_connection()
    cursor = conn.cursor()

    # Collect baseline row count and file size before optimization.
    cursor.execute("SELECT COUNT(*) FROM player_rankings")
    before_count = cursor.fetchone()[0]
    before_size = os.path.getsize(DB_FILE) if os.path.exists(DB_FILE) else 0

    # Enumerate all (player_id, mode) pairs to process.
    cursor.execute("SELECT DISTINCT player_id, mode FROM player_rankings")
    player_mode_pairs = cursor.fetchall()

    total_pairs = len(player_mode_pairs)
    if total_pairs == 0:
        logger.info("数据库为空，无需优化")
        return

    def _is_valid_lv(value) -> bool:
        try:
            return value is not None and int(value) > 0
        except Exception:
            return False

    # Process in one transaction for better performance.
    conn.execute("BEGIN TRANSACTION")
    deleted_total = 0
    delete_batch_size = 500
    lv_repaired_total = 0

    # Optional progress visualization.
    iterator = tqdm(player_mode_pairs, desc="Optimize DB", unit="pair") if HAS_TQDM else player_mode_pairs

    try:
        for player_id, mode in iterator:
            # Load all records for this player/mode in ascending time order.
            cursor.execute('''
                SELECT id, rank, name, lv, exp, acc, combo, pc, crawl_time
                FROM player_rankings
                WHERE player_id = ? AND mode = ?
                ORDER BY crawl_time ASC, id ASC
            ''', (player_id, mode))
            rows = cursor.fetchall()

            if len(rows) <= 1:
                continue

            to_delete = []
            i = 0
            while i < len(rows):
                j = i + 1
                # Find a contiguous segment of identical ranking payload.
                # Align with save_player_ranking_record current_core:
                # (rank, name, exp, acc, combo, pc) -- lv is intentionally excluded.
                while j < len(rows) and (
                    rows[j][1], rows[j][2], rows[j][4], rows[j][5], rows[j][6], rows[j][7]
                ) == (
                    rows[i][1], rows[i][2], rows[i][4], rows[i][5], rows[i][6], rows[i][7]
                ):
                    j += 1

                # Extra handling for legacy lv bug (0 <-> actual oscillation):
                # if a stable core-segment contains any valid non-zero lv, normalize
                # kept boundary rows' lv to the segment's latest non-zero lv.
                preferred_lv = None
                for k in range(j - 1, i - 1, -1):
                    candidate_lv = rows[k][3]
                    if _is_valid_lv(candidate_lv):
                        preferred_lv = int(candidate_lv)
                        break

                if preferred_lv is not None:
                    start_id, start_lv = rows[i][0], rows[i][3]
                    end_id, end_lv = rows[j - 1][0], rows[j - 1][3]

                    if not _is_valid_lv(start_lv):
                        cursor.execute("UPDATE player_rankings SET lv = ? WHERE id = ?", (preferred_lv, start_id))
                        lv_repaired_total += 1

                    if end_id != start_id and not _is_valid_lv(end_lv):
                        cursor.execute("UPDATE player_rankings SET lv = ? WHERE id = ?", (preferred_lv, end_id))
                        lv_repaired_total += 1

                # For [i, j-1] with length>1, drop middle rows and keep boundaries.
                if j - i > 1:
                    for k in range(i + 1, j - 1):
                        to_delete.append(rows[k][0])
                i = j

            if to_delete:
                for start in range(0, len(to_delete), delete_batch_size):
                    chunk = to_delete[start:start + delete_batch_size]
                    placeholders = ','.join('?' * len(chunk))
                    cursor.execute(f"DELETE FROM player_rankings WHERE id IN ({placeholders})", chunk)
                deleted_total += len(to_delete)
                if HAS_TQDM:
                    _ = iterator.set_postfix(deleted=deleted_total)
    except Exception:
        conn.rollback()
        raise

    conn.commit()

    # Collect row count and file size after optimization.
    cursor.execute("SELECT COUNT(*) FROM player_rankings")
    after_count = cursor.fetchone()[0]
    # Run VACUUM to reclaim freed database space.
    conn.execute("VACUUM")
    after_size = os.path.getsize(DB_FILE)

    # Compute optimization deltas and ratios.
    delta_count = before_count - after_count
    delta_size = before_size - after_size
    count_ratio = (delta_count / before_count * 100) if before_count > 0 else 0
    size_ratio = (delta_size / before_size * 100) if before_size > 0 else 0

    # Print optimization report.
    print("\n" + "="*50)
    print("数据库优化完成")
    print("="*50)
    print(f"优化前记录数: {before_count}")
    print(f"优化后记录数: {after_count}")
    print(f"删除记录数: {delta_count} ({count_ratio:.2f}%)")
    print(f"修复LV异常: {lv_repaired_total} 条边界记录")
    print(f"优化前文件大小: {before_size/1024:.2f} KB")
    print(f"优化后文件大小: {after_size/1024:.2f} KB")
    print(f"节省空间: {delta_size/1024:.2f} KB ({size_ratio:.2f}%)")
    print("="*50)

    logger.info(
        "数据库优化完成: 删除 %d 条记录，修复LV异常 %d 条边界记录，节省 %.2f KB",
        delta_count,
        lv_repaired_total,
        delta_size / 1024,
    )
    return delta_count


def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(description="Malody rankings crawler")
    _ = parser.add_argument('--all', action='store_true',
                            help='Crawl leaderboard and player pages, including leaderboard players into queue')
    _ = parser.add_argument('--leaderboard-only', action='store_true',
                            help='Only crawl leaderboard, skip all player-profile crawling')
    _ = parser.add_argument('--players-only', action='store_true',
                            help='Only crawl player profiles from players.txt, skip leaderboard')
    _ = parser.add_argument('--no-player-crawl', action='store_true',
                            help='Disable default player-profile crawl')
    _ = parser.add_argument('--migrate-db', action='store_true',
                            help='Run database migration tasks')
    _ = parser.add_argument('--once', action='store_true',
                            help='Run one crawl cycle and exit')
    _ = parser.add_argument('--save-excel', action='store_true',
                            help='Save output to Excel files')
    _ = parser.add_argument('--push-to-git', action='store_true',
                            help='Push updated data to Git repository')
    _ = parser.add_argument('--optimize-db', action='store_true',
                            help='Optimize database and exit')
    _ = parser.add_argument('--mm-sync', action='store_true',
                            help='Run MM ranking and MMR sync in normal cycle')
    _ = parser.add_argument('--mm-only', action='store_true',
                            help='Run only MM ranking and MMR sync')
    _ = parser.add_argument('--mm-limit', type=int, default=DEFAULT_MM_LIMIT,
                            help=f'Max rows per mode for MM sync (default {DEFAULT_MM_LIMIT})')
    _ = parser.add_argument('--skip-mm-ranking', action='store_true',
                            help='Skip MM ranking crawl and only run MMR sync')
    _ = parser.add_argument('--skip-mmr', action='store_true',
                            help='Skip MMR sync and only run MM ranking crawl')

    _ = parser.add_argument('--ranking-source', choices=['page', 'newapi'], default='page',
                            help='Ranking source: page=legacy HTML parse, newapi=/ranking/global')
    _ = parser.add_argument('--ranking-limit', type=int, default=DEFAULT_MM_LIMIT,
                            help=f'Max rows per mode when ranking-source=newapi (default {DEFAULT_MM_LIMIT})')

    return parser.parse_args()

def run_players_only():
    """Run player-profile crawling only (no leaderboard crawling)."""
    init_database()
    config_players = load_player_config()
    if config_players:
        add_players_to_queue(config_players)
        run_player_crawler()
    else:
        logger.info("没有配置玩家，跳过爬取")
    DatabaseManager().close_connection()


def main():
    args = parse_arguments()

    if args.migrate_db:
        migrate_database()
        return

    init_database()

    # Preload config players once before first cycle; later cycles reload again.
    if not args.mm_only:
        config_players = load_player_config()
        if config_players:
            add_players_to_queue(config_players)
            logger.info("Loaded %d players from config file.", len(config_players))

    if args.optimize_db:
        optimize_database()
        DatabaseManager().close_connection()
        return

    if args.players_only:
        run_players_only()
        return

    if args.mm_only:
        mm_limit = max(1, int(args.mm_limit or DEFAULT_MM_LIMIT))
        logger.info(
            "Run mm-only cycle: mm_limit=%d skip_mm_ranking=%s skip_mmr=%s",
            mm_limit,
            args.skip_mm_ranking,
            args.skip_mmr,
        )
        try:
            run_mm_sync_cycle(
                mm_limit=mm_limit,
                include_mm_ranking=not args.skip_mm_ranking,
                include_mmr=not args.skip_mmr,
            )
        finally:
            DatabaseManager().close_connection()
        return

    # Decide whether player crawling is enabled for this run.
    crawl_players = True
    if args.no_player_crawl or args.leaderboard_only:
        crawl_players = False

    # Decide whether to enqueue leaderboard players (only with --all).
    crawl_leaderboard_players = args.all

    # Decide whether to save Excel snapshots and push Git updates.
    save_excel = args.save_excel
    push_to_git = args.push_to_git
    mm_sync = args.mm_sync
    mm_limit = max(1, int(args.mm_limit or DEFAULT_MM_LIMIT))
    ranking_source = args.ranking_source
    ranking_limit = max(1, int(args.ranking_limit or DEFAULT_MM_LIMIT))

    if args.once:
        if args.leaderboard_only:
            crawl_players = False
            crawl_leaderboard_players = False
        if args.all:
            crawl_players = True
            crawl_leaderboard_players = True
            if args.no_player_crawl:
                logger.warning("同时指定了 --all 和 --no-player-crawl，以 --all 为准，将爬取玩家")
                crawl_players = True
                crawl_leaderboard_players = True

        run_crawler_cycle(
            crawl_players=crawl_players,
            crawl_leaderboard_players=crawl_leaderboard_players,
            save_excel=save_excel,
            push_to_git=push_to_git,
            mm_sync=mm_sync,
            mm_limit=mm_limit,
            skip_mm_ranking=args.skip_mm_ranking,
            skip_mmr=args.skip_mmr,
            ranking_source=ranking_source,
            ranking_limit=ranking_limit,
        )
        DatabaseManager().close_connection()
        return

    try:
        while True:
            with stop_lock:
                if stop_requested:
                    logger.info("程序被终止")
                    break

            try:
                run_crawler_cycle(
                    crawl_players=crawl_players,
                    crawl_leaderboard_players=crawl_leaderboard_players,
                    save_excel=save_excel,
                    push_to_git=push_to_git,
                    mm_sync=mm_sync,
                    mm_limit=mm_limit,
                    skip_mm_ranking=args.skip_mm_ranking,
                    skip_mmr=args.skip_mmr,
                    ranking_source=ranking_source,
                    ranking_limit=ranking_limit,
                )
            except Exception as e:
                logger.exception("主循环发生未处理异常")

            logger.info("等待30分钟后重启...")

            for i in range(30):
                with stop_lock:
                    if stop_requested:
                        logger.info("程序被终止")
                        break
                time.sleep(60)
                gc.collect()
    finally:
        DatabaseManager().close_connection()


if __name__ == "__main__":
    main()
